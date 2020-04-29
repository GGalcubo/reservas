# -*- coding: utf-8 -*-
from __future__ import unicode_literals
# master
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Permission
from django.template.defaulttags import register
from django.conf import settings
from .models import *
from django.http import HttpResponse, JsonResponse
from django.core import serializers
from datetime import timedelta
import json
import os
import datetime

@login_required
def dashboard(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	mensaje = ""
	context = { 'mensaje':mensaje }
	return render(request, 'sistema/dashboard.html', context)

@login_required
def operaciones(request):
	permiso = obtenerPermiso(request)
	if not validarUrlPorRol(request):
		if 'unidades' in permiso:
			return redirect("/sistema/asignaciones/")
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	mensaje = ""

	if 'unidades' in permiso:
		viajes = []
		viajesQ = Viaje.objects.filter(estado_id__in=[4,5,6])
		for v in viajesQ:
			if validaViajeUnidad(request, v):
				viajes.append(v)
		return redirect("/asignaciones/")
	else:
		viajes = Viaje.objects.all()
	estados = Estado.objects.all()
	categoria_viajes = CategoriaViaje.objects.all()
	#unidades = Unidad.objects.extra(select={'id_fake': 'CAST(id_fake AS INTEGER)'}).order_by('id_fake')
	unidades = Unidad.objects.filter(baja=False).order_by('id_fake')

	context = {'mensaje': mensaje, 'estados': estados,'categoria_viajes': categoria_viajes, 'unidades': unidades,'permiso':permiso}
	return render(request, 'sistema/operaciones.html', context)

@login_required
def asignaciones(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	mensaje = ""
	permiso = obtenerPermiso(request)
	#viajes = Viaje.objects.all()
	#unidades = Unidad.objects.all()
	estados = Estado.objects.filter(id__in=[3, 4, 5, 6, 7, 8, 9, 10])
	#'viajes': viajes, 'unidades': unidades,

	context = {'mensaje':mensaje, 'estados': estados,'permiso':permiso }
	return render(request, 'sistema/asignaciones.html', context)

@login_required
def pasajeros(request):

    mensaje = ""
    permiso = obtenerPermiso(request)
    id_viaje = request.GET.get('idViaje', "")
    viaje = Viaje.objects.get(id=id_viaje)

    context = {'mensaje':mensaje,'viaje':viaje,'permiso':permiso}
    return render(request, 'sistema/pasajeros.html', context)

@login_required
def getViajesAsignacionesPorFecha(request):
	mensaje = ""
	date = getAAAAMMDD(request.POST.get('date', False))
	unidades = getUnidadesByUser(request)
	estados_get_seleccionados = request.POST.getlist('estados_selecionados[]', False)
	estados_seleccionados = []
	if estados_get_seleccionados == False:
		estados_seleccionados = [3, 4]
	else:
		for i in estados_get_seleccionados:
			estados_seleccionados.append(i)

	viajes = Viaje.objects.filter(fecha=date).filter(estado__in=estados_seleccionados).filter(unidad__in=unidades)
	context = {'mensaje': mensaje, 'viajes': viajes}
	return render(request, 'sistema/grillaViajesAsignaciones.html', context)

@login_required
def buscarViajes(request):
	mensaje = ""

	unidad          = request.POST.get('unidad', False)
	cliente         = request.POST.get('cliente', False)
	pasajero        = request.POST.get('pasajero', False)
	solicitante     = request.POST.get('solicitante', False)
	centroDeCosto   = request.POST.get('centroDeCosto', False)
	estados 		= request.POST.get('estados', False)
	desde           = request.POST.get('desde', False)
	hasta           = request.POST.get('hasta', False)

	print estados

	estList = []
	if estados != "":
		for c in estados.split(","):
			if c != "":
				estList.append(int(c))

	viajes = Viaje.objects.filter(fecha__gte=desde, fecha__lte=hasta)

	if pasajero:
		viajes = viajes.filter(pasajero_id=pasajero)

	if solicitante:
		viajes = viajes.filter(solicitante_id=solicitante)

	if centroDeCosto:
		viajes = viajes.filter(centro_costo_id=centroDeCosto)

	if unidad:
		viajes = viajes.filter(unidad_id=unidad)

	if cliente:
		viajes = viajes.filter(cliente_id=cliente)

	if estList:
		viajes = viajes.filter(estado_id__in=estList)

	context = {'mensaje': mensaje, 'viajes': viajes}
	return render(request, 'sistema/grillaViajesExportar.html', context)

@login_required
def getClientes(request):
    clientes = serializers.serialize('json', Cliente.objects.filter(baja=False))
    return HttpResponse(clientes, content_type='application/json')

@login_required
def getClienteById(request):
    cliente = Cliente.objects.get(id=request.POST.get('cliente_id', False))
    pasajero_id = request.POST.get('pasajero', 0)
    if pasajero_id == '':
        pasajero_id = 0
    personacliente = []
    for i in cliente.personacliente_set.all():
        if i.persona.baja is False:
            if ((i.persona.tipo_persona.tipo_persona == 'Solicitante') or (int(i.persona.id) == int(pasajero_id))):
                personacliente.append({
                    'id':i.persona.id,
                    'nombre':i.persona.apellido + ' ' + i.persona.nombre,
                    'tipo_persona':i.persona.tipo_persona.tipo_persona,
                    'telefono':i.persona.telefono
                })
    personacliente = sorted(personacliente, key = lambda  i: (i['tipo_persona'], i['nombre'].lower()))


    centrocosto = []
    for c in cliente.centrocosto_set.filter(baja=False):
        centrocosto.append({'id':c.id,'nombre':c.nombre})

    data = {
        'id': cliente.id,
        'calle': cliente.calle,
        'altura': cliente.altura,
        'piso': cliente.piso,
        'depto': cliente.depto,
        'moroso': cliente.moroso,
        'razon_social': cliente.razon_social,
        'telefono': cliente.telefonoPrincipal(),
        'centro_costos': list(centrocosto),
        'personascliente': list(personacliente)
    }

    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')


@login_required
def getPasajerosByClienteId(request):
    cliente = Cliente.objects.get(id=request.POST.get('cliente_id', False))
    pasajero_id = request.POST.get('pasajero', 0)
    if pasajero_id == '':
        pasajero_id = 0
    personacliente = []
    for i in cliente.personacliente_set.all():
        if i.persona.baja is False:
            if (i.persona.tipo_persona.tipo_persona == 'Pasajero' and i.persona.pasajero_frecuente == 1):
                personacliente.append({
                    'id':i.persona.id,
                    'nombre':i.persona.apellido + ' ' + i.persona.nombre,
                    'tipo_persona':i.persona.tipo_persona.tipo_persona,
                    #'telefono':i.persona.telefono
                })
    personacliente = sorted(personacliente, key = lambda i: (i['nombre'].lower()))


    data = {
        'id': cliente.id,
        'personascliente': list(personacliente)
    }

    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')


@login_required
def altaViaje(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	mensaje     = ""
	es_nuevo    = 1
	viaje       = Viaje()
	viaje.id    = 0

	context = {'mensaje': mensaje,
				#'clientes':Cliente.objects.filter(baja=False),
				'tipoobservacion':TipoObservacion.objects.all(),
				'tipo_pago':TipoPagoViaje.objects.all(),
				'unidades':Unidad.objects.filter(baja=False).values_list('id', 'id_fake', 'identificacion').order_by('id_fake'),
				'estados':Estado.objects.all(),
				'categoria_viajes':CategoriaViaje.objects.all(),
				'tarifarios':Tarifario.objects.filter(baja=False),
				#'destinos':TrayectoDestino.objects.all(),
				'provincias':Provincia.objects.all(),
				'es_nuevo':es_nuevo,
				'viaje':viaje}

	return render(request, 'sistema/viaje.html', context)

@login_required
def editaViaje(request):
	mensaje     = ""
	es_nuevo    = 0
	id_viaje    = request.GET.get('idViaje', "")
	is_clone    = request.GET.get('c', "")
	if request.GET.get('msg', "") == '1':
		mensaje = 'El viaje se creo correctamente.'
	viaje = Viaje.objects.get(id=id_viaje)
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

 
	if not validarViajeUsuarioUnidad(request, viaje):
		mensaje = ""
		context = { 'mensaje':mensaje }

		return render(request, 'sistema/urlBloqueada.html', context)
	context = {'mensaje': mensaje,
				#'clientes':Cliente.objects.filter(baja=False),
				'tipoobservacion':TipoObservacion.objects.all(),
				'is_clone':is_clone,
				'tipo_pago':TipoPagoViaje.objects.all(),
				#'itemsviaje':ItemViaje.objects.filter(viaje_id=id_viaje),
				'unidades':Unidad.objects.filter(baja=False).values_list('id', 'id_fake', 'identificacion').order_by('id_fake'),
				'estados':Estado.objects.all(),
				'categoria_viajes':CategoriaViaje.objects.all(),
				'tarifarios':Tarifario.objects.filter(baja=False),
				#'destinos':TrayectoDestino.objects.all(),
				#'localidades':Localidad.objects.filter(baja=False),
				'provincias':Provincia.objects.all(),
				'es_nuevo':es_nuevo,
				'viaje':viaje,
                'permiso':obtenerPermiso(request),
				'trayectos':Trayecto.objects.filter(viaje_id=id_viaje)}
	return render(request, 'sistema/viaje.html', context)


@login_required
def clonarViaje(request):
    viaje_a_clonar = Viaje.objects.get(id=request.POST.get('idViaje', ""))
    viaje_clonado = Viaje()
    viaje_clonado.centro_costo = viaje_a_clonar.centro_costo
    viaje_clonado.cliente = viaje_a_clonar.cliente
    viaje_clonado.solicitante = viaje_a_clonar.solicitante
    viaje_clonado.pasajero = viaje_a_clonar.pasajero
    viaje_clonado.Cod_ext_viaje = viaje_a_clonar.Cod_ext_viaje
    viaje_clonado.nro_aux = viaje_a_clonar.nro_aux
    viaje_clonado.categoria_viaje = viaje_a_clonar.categoria_viaje
    #viaje_clonado.unidad = viaje_a_clonar.unidad
    viaje_clonado.creadofecha = fecha()
    viaje_clonado.creadopor = request.user
    viaje_clonado.save()

    trayectos_a_clonar = Trayecto.objects.filter(viaje_id=viaje_a_clonar.id)
    for t in trayectos_a_clonar:
        trayectos_clonados = Trayecto()
        trayectos_clonados.viaje_id = viaje_clonado.id
        trayectos_clonados.pasajero = t.pasajero
        trayectos_clonados.altura_desde = t.altura_desde
        trayectos_clonados.altura_hasta = t.altura_hasta
        trayectos_clonados.calle_desde = t.calle_desde
        trayectos_clonados.calle_hasta = t.calle_hasta
        trayectos_clonados.comentario = t.comentario
        trayectos_clonados.compania_desde = t.compania_desde
        trayectos_clonados.compania_hasta = t.compania_hasta
        trayectos_clonados.destino_desde = t.destino_desde
        trayectos_clonados.destino_hasta = t.destino_hasta
        trayectos_clonados.entre_desde = t.entre_desde
        trayectos_clonados.entre_hasta = t.entre_hasta
        trayectos_clonados.localidad_desde = t.localidad_desde
        trayectos_clonados.localidad_hasta = t.localidad_hasta
        trayectos_clonados.provincia_desde = t.provincia_desde
        trayectos_clonados.provincia_hasta = t.provincia_hasta
        trayectos_clonados.tramoppalflag = t.tramoppalflag
        trayectos_clonados.vuelo_desde = t.vuelo_desde
        trayectos_clonados.vuelo_hasta = t.vuelo_hasta
        trayectos_clonados.save()
    data = {
        'error': '0',
        'viaje_a_clonar': viaje_clonado.id
    }
    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')



@login_required
def cambiarUnidadViaje(request):
    viaje = Viaje.objects.get(id=request.POST.get('idViaje', False))
    unidad = request.POST.get('unidad_id', '')
    if unidad != '':
        viaje.unidad = Unidad.objects.get(id=unidad)

    viaje.save()

    data = {
        'error': '0',
        'msg': 'La unidad ha sido modificada'
    }
    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')


@login_required
def getDatosUnidad(request):
    unidad = ''
    if request.POST.get('unidad_id', False) != '':
        unidad = Unidad.objects.get(id=request.POST.get('unidad_id', False))

    if unidad:
        data = {
            'identificacion': unidad.identificacion if unidad.identificacion else '',
            'telefono': unidad.telefono if unidad.telefono else '',
            'mail': unidad.mail if unidad.mail else '',
            'marca': unidad.vehiculo.marca if unidad.vehiculo and unidad.vehiculo.marca else '',
            'modelo': unidad.vehiculo.modelo if unidad.vehiculo and unidad.vehiculo.modelo else '',
            'color': unidad.vehiculo.color if unidad.vehiculo and unidad.vehiculo.color else '',
            'patente': unidad.vehiculo.patente if unidad.vehiculo and unidad.vehiculo.patente else ''
        }
    else:
        data = {}
    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')

@login_required
def guardarViaje(request):
    es_nuevo = request.POST.get('es_nuevo', "1")
    mensaje = ''

    cliente = Cliente.objects.get(id=request.POST.get('cliente', False))
    if cliente.moroso == True:
        mensaje = 'Comunicarse con administración, cliente inhabilitado.'
        data = {
            'error': '1',
            'msg': mensaje
        }
        dump = json.dumps(data)
        return HttpResponse(dump, content_type='application/json')

    if es_nuevo == "1":
        viaje = Viaje()
        viaje.creadofecha = fecha()
        viaje.creadopor = request.user
    else:
        viaje = Viaje.objects.get(id=request.POST.get('idViaje', False))
        mensaje = 'El viaje se actualizo correctamente.'

    viaje.estado 				= Estado.objects.get(id=request.POST.get('estado', False))
    viaje.cliente 				= cliente
    viaje.categoria_viaje 		= CategoriaViaje.objects.get(id=request.POST.get('categoria_viaje', False))
    viaje.solicitante 			= Persona.objects.get(id=request.POST.get('contacto', False))
    viaje.centro_costo 			= CentroCosto.objects.get(id=request.POST.get('centro_costos', False))
    #pasajero                   = Persona.objects.get(id=request.POST.get('pasajero', False))
    #viaje.pasajero 				= pasajero
    fecha_tmp 					= request.POST.get('fecha', "")
    viaje.fecha 				= fecha_tmp[6:10] + fecha_tmp[3:5] + fecha_tmp[0:2]
    viaje.hora 					= request.POST.get('hora', "")
    #viaje.hora_estimada 		= request.POST.get('hora_estimada', "")
    viaje.tarifapasada 			= request.POST.get('tarifa_pasada', "")
    viaje.Cod_ext_viaje         = request.POST.get('cod_externo', "")
    viaje.nro_aux               = request.POST.get('nro_aux', "")
    viaje.nropasajeros          = request.POST.get('pasajero_cant', "")
    viaje.tipo_pago             = TipoPagoViaje.objects.get(id=request.POST.get('tipo_pago', False))

    if viaje.cabecera is False:
        if request.POST.get('bilingue', False) == '':
            viaje.bilingue = False
        else:
            viaje.bilingue = True
        viaje.maletas               = request.POST.get('maletas', "")
        viaje.espera                = request.POST.get('tiempo_espera', "")
        viaje.dispo                 = request.POST.get('tiempo_hs_dispo', "")
        viaje.peajes                = request.POST.get('peaje', "")
        viaje.parking               = request.POST.get('estacionamiento', "")
        viaje.otro                  = request.POST.get('otros', "")

    unidad 						= request.POST.get('unidad', '')
    if unidad != '':
        viaje.unidad 			= Unidad.objects.get(id=unidad)
    data = {
        'error': '0',
        'msg': mensaje
    }

    if viaje.fecha > viaje.centro_costo.fecha_fin:
        data = {
            'error': '1',
            'msg': 'La fecha del viaje es posterior al vencimiento del centro de costos'
        }
        dump = json.dumps(data)
        return HttpResponse(dump, content_type='application/json')

    viaje.save()

    guardarTrayecto(request, True, viaje.id)

    guardarHistorial(viaje, 'estado', viaje.estado.estado, request.user)
    guardarHistorial(viaje, 'cliente', viaje.cliente.razon_social, request.user)
    guardarHistorial(viaje, 'categoria viaje', viaje.categoria_viaje.categoria, request.user)
    guardarHistorial(viaje, 'solicitante', viaje.solicitante.apellido + ' ' + viaje.solicitante.nombre, request.user)
    guardarHistorial(viaje, 'centro costo', viaje.centro_costo.nombre, request.user)
    #guardarHistorial(viaje, 'pasajero', viaje.pasajero.apellido + ' ' + viaje.pasajero.nombre, request.user)
    guardarHistorial(viaje, 'hora', viaje.hora, request.user)
    #guardarHistorial(viaje, 'hora estimada', viaje.hora_estimada, request.user)
    guardarHistorial(viaje, 'tarifa pasada', viaje.tarifapasada, request.user)
    guardarHistorial(viaje, 'cod ext viaje', viaje.Cod_ext_viaje, request.user)
    guardarHistorial(viaje, 'nro aux', viaje.nro_aux, request.user)
    guardarHistorial(viaje, 'tipo pago', viaje.tipo_pago.tipo_pago_viaje, request.user)

    ##guardarHistorial(viaje, 'bilingue', viaje.bilingue, request.user)
    guardarHistorial(viaje, 'pasajero cant', viaje.nropasajeros, request.user)
    guardarHistorial(viaje, 'maletas', viaje.maletas, request.user)
    guardarHistorial(viaje, 'espera', viaje.espera, request.user)
    guardarHistorial(viaje, 'hora disponible', viaje.dispo, request.user)
    guardarHistorial(viaje, 'peajes', viaje.peajes, request.user)
    guardarHistorial(viaje, 'parking', viaje.parking, request.user)
    guardarHistorial(viaje, 'otro', viaje.otro, request.user)
    if unidad != '':
        guardarHistorial(viaje, 'unidad', viaje.unidad.id_fake + ' ' + viaje.unidad.identificacion, request.user)

    if viaje.getTrayectoPrincipal() != '':
        trayecto = viaje.getTrayectoPrincipal()
        #trayecto.pasajero = pasajero
        trayecto.save()

    #guardaViajePasajero(pasajero, True, viaje)
    guardaObsViaje(request.POST.get('comentario_chofer', ''), viaje, request)
    guardarHistorial(viaje, 'comentario chofer', request.POST.get('comentario_chofer', ''), request.user)
    guardaItemViaje(request.POST.get('importe_efectivo', ''), 12, 1, viaje, False)
    guardarHistorial(viaje, 'importe efectivo', request.POST.get('importe_efectivo', ''), request.user)

    if viaje.estado.id == 7 and viaje.calculo_admin is False:

        guardaItemViajeCostoProveedor('', 8, 1, viaje, False)
        guardaItemViajeCostoCliente('', 2, 1, viaje, False)

        item_viaje_hs_dispo = guardaItemViajeHsDispo('', 18, request.POST.get('tiempo_hs_dispo', ''), viaje, False)
        item_viaje_hs_dispo_admin = guardaItemViajeHsDispoAdmin('', 13, request.POST.get('tiempo_hs_dispo', ''), viaje, False)

        item_viaje_espera = guardaItemViajeEspera('', 1, request.POST.get('tiempo_espera', ''), viaje, False)
        item_viaje_espera_admin = guardaItemViajeEsperaAdmin('', 14, request.POST.get('tiempo_espera', ''), viaje, False)

        guardaItemViajeBilingue('', 3, request.POST.get('bilingue', ''), viaje, False, item_viaje_espera, item_viaje_hs_dispo)
        guardaItemViajeBilingueAdmin('', 9, request.POST.get('bilingue', ''), viaje, False, item_viaje_espera_admin, item_viaje_hs_dispo_admin)

        guardaItemViajeMaletas('', 4, request.POST.get('maletas', ''), viaje, False)

        guardaItemViaje              (request.POST.get('otros', ''), 17, 1, viaje, False)
        guardaItemViaje              (request.POST.get('peaje', ''), 6, 1, viaje, False)
        guardaItemViaje              (request.POST.get('estacionamiento', ''), 5, 1, viaje, False)

        guardarHistorial(viaje, 'otros', request.POST.get('otros', ''), request.user)
        guardarHistorial(viaje, 'peaje', request.POST.get('peaje', ''), request.user)
        guardarHistorial(viaje, 'estacionamiento', request.POST.get('estacionamiento', ''), request.user)
        guardarHistorial(viaje, 'comentario_chofer', request.POST.get('comentario_chofer', ''), request.user)
        guardarHistorial(viaje, 'tiempo_hs_dispo', request.POST.get('tiempo_hs_dispo', ''), request.user)
        guardarHistorial(viaje, 'tiempo_espera', request.POST.get('tiempo_espera', ''), request.user)
        guardarHistorial(viaje, 'bilingue', request.POST.get('bilingue', ''), request.user)
        guardarHistorial(viaje, 'maletas', request.POST.get('maletas', ''), request.user)

        viaje.calculo_admin = True
        viaje.cabecera = True
        viaje.save()

        items_viaje = serializers.serialize('json', ItemViaje.objects.filter(viaje=viaje))
        return HttpResponse(items_viaje, content_type='application/json')

    if es_nuevo == "1":
        data = {
            'url': '/sistema/editaViaje/?idViaje=' + str(viaje.id) + '&msg=1',
            'viaje': viaje.id
        }

    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')

def guardarHistorial(viaje, field, value, edit_by):
    if value != '':
        historial = ViajeHistorial.objects.filter(viaje=viaje, campo_modificado=field).last()
        if historial != None:
            if historial.valor_actual != value:
                historial_nuevo = ViajeHistorial()
                historial_nuevo.viaje = viaje
                historial_nuevo.campo_modificado = field
                historial_nuevo.valor_anterior = historial.valor_actual
                historial_nuevo.valor_actual = value
                historial_nuevo.fecha = fecha()
                historial_nuevo.usuario = edit_by
                historial_nuevo.save()
        else:
            historial = ViajeHistorial()
            historial.viaje = viaje
            historial.campo_modificado = field
            historial.valor_anterior = ''
            historial.valor_actual = value
            historial.fecha = fecha()
            historial.usuario = edit_by
            historial.save()

def guardaViajeAdmin(request):
    viaje = Viaje.objects.get(id=request.POST.get('viaje', False))
    manual = False
    if request.POST.get('metodo', False) == 'guarda':
        manual = True

    guardaItemViaje(request.POST.get('admin_otros_proveedor', ''), 16, 1, viaje, manual)
    guardaItemViaje(request.POST.get('admin_otros_cliente', ''), 17, 1, viaje, manual)
    guardarHistorial(viaje, 'admin_otros_proveedor', request.POST.get('admin_otros_proveedor', ''), request.user)
    guardarHistorial(viaje, 'admin_otros_cliente', request.POST.get('admin_otros_cliente', ''), request.user)

    guardaItemViaje(request.POST.get('admin_peaje_proveedor', ''), 15, 1, viaje, manual)
    guardaItemViaje(request.POST.get('admin_peaje_cliente', ''), 6, 1, viaje, manual)
    guardarHistorial(viaje, 'admin_peaje_proveedor', request.POST.get('admin_peaje_proveedor', ''), request.user)
    guardarHistorial(viaje, 'admin_peaje_cliente', request.POST.get('admin_peaje_cliente', ''), request.user)

    guardaItemViaje(request.POST.get('admin_estacionamiento_proveedor', ''), 11, 1, viaje, manual)
    guardaItemViaje(request.POST.get('admin_estacionamiento_cliente', ''), 5, 1, viaje, manual)
    guardarHistorial(viaje, 'admin_estacionamiento_proveedor', request.POST.get('admin_peaje_cliente', ''), request.user)
    guardarHistorial(viaje, 'admin_estacionamiento_cliente', request.POST.get('admin_estacionamiento_cliente', ''), request.user)

    guardaItemViajeCostoProveedor(request.POST.get('admin_costo_proveedor', ''), 8, 1, viaje, manual)
    guardaItemViajeCostoCliente(request.POST.get('admin_costo_cliente', ''), 2, 1, viaje, manual)
    guardarHistorial(viaje, 'admin_costo_proveedor', request.POST.get('admin_costo_proveedor', ''), request.user)
    guardarHistorial(viaje, 'admin_costo_cliente', request.POST.get('admin_costo_cliente', ''), request.user)

    item_viaje_hs_dispo = guardaItemViajeHsDispo(request.POST.get('admin_hs_dispo_cliente', ''), 18, request.POST.get('admin_cant_hs_dispo_cliente', ''), viaje, manual)
    item_viaje_hs_dispo_admin = guardaItemViajeHsDispoAdmin(request.POST.get('admin_hs_dispo_proveedor', ''), 13, request.POST.get('admin_cant_hs_dispo_proveedor', ''), viaje, manual)
    guardarHistorial(viaje, 'admin_hs_dispo_cliente', request.POST.get('admin_hs_dispo_cliente', ''), request.user)
    guardarHistorial(viaje, 'admin_cant_hs_dispo_cliente', request.POST.get('admin_cant_hs_dispo_cliente', ''), request.user)
    guardarHistorial(viaje, 'admin_hs_dispo_proveedor', request.POST.get('admin_hs_dispo_proveedor', ''), request.user)
    guardarHistorial(viaje, 'admin_cant_hs_dispo_proveedor', request.POST.get('admin_cant_hs_dispo_proveedor', ''), request.user)

    item_viaje_espera = guardaItemViajeEspera(request.POST.get('admin_espera_cliente', ''), 1, request.POST.get('admin_cant_espera_cliente', ''), viaje, manual)
    item_viaje_espera_admin = guardaItemViajeEsperaAdmin(request.POST.get('admin_espera_proveedor', ''), 14, request.POST.get('admin_cant_espera_proveedor', ''), viaje, manual)
    guardarHistorial(viaje, 'admin_espera_cliente', request.POST.get('admin_espera_cliente', ''), request.user)
    guardarHistorial(viaje, 'admin_cant_espera_cliente', request.POST.get('admin_cant_espera_cliente', ''), request.user)
    guardarHistorial(viaje, 'admin_espera_proveedor', request.POST.get('admin_espera_proveedor', ''), request.user)
    guardarHistorial(viaje, 'admin_cant_espera_proveedor', request.POST.get('admin_cant_espera_proveedor', ''), request.user)

    guardaItemViajeBilingue(request.POST.get('admin_costo_bilingue_cliente', ''), 3, request.POST.get('admin_bilingue_cliente', ''), viaje, manual, item_viaje_espera, item_viaje_hs_dispo)
    guardaItemViajeBilingueAdmin(request.POST.get('admin_costo_bilingue_proveedor', ''), 9, request.POST.get('admin_bilingue_proveedor', ''), viaje, manual, item_viaje_espera_admin, item_viaje_hs_dispo_admin)
    guardarHistorial(viaje, 'admin_costo_bilingue_cliente', request.POST.get('admin_costo_bilingue_cliente', ''), request.user)
    guardarHistorial(viaje, 'admin_costo_bilingue_proveedor', request.POST.get('admin_costo_bilingue_proveedor', ''), request.user)

    guardaItemViajeMaletas(request.POST.get('admin_costo_maletas_cliente', ''), 4, request.POST.get('admin_cant_maletas_cliente', ''), viaje, manual)
    guardaItemViajeMaletasAdmin(request.POST.get('admin_costo_maletas_proveedor', ''), 10, request.POST.get('admin_cant_maletas_proveedor', ''), viaje, manual)
    guardarHistorial(viaje, 'admin_costo_maletas_cliente', request.POST.get('admin_costo_bilingue_proveedor', ''), request.user)
    guardarHistorial(viaje, 'admin_cant_maletas_cliente', request.POST.get('admin_cant_maletas_cliente', ''), request.user)
    guardarHistorial(viaje, 'admin_costo_maletas_proveedor', request.POST.get('admin_costo_maletas_proveedor', ''), request.user)
    guardarHistorial(viaje, 'admin_cant_maletas_proveedor', request.POST.get('admin_cant_maletas_proveedor', ''), request.user)

    items_viaje = serializers.serialize('json', ItemViaje.objects.filter(viaje=viaje))
    return HttpResponse(items_viaje, content_type='application/json')

def validateGetTarifaTrayecto(tarifario, viaje):
    try:
        if tarifario == 'centro_costo':
            tarifario = viaje.centro_costo.tarifario
            #print viaje.categoria_viaje_id
            #print tarifario.id
            #print viaje.getTrayectoPrincipal().localidad_desde.id
            #print viaje.getTrayectoPrincipal().localidad_hasta.id
        elif tarifario == 'unidad':
            tarifario = viaje.unidad.tarifario
    except Exception as e:
        tarifario = ''
    try:
        base = getTarifaTrayecto(viaje.categoria_viaje_id, tarifario, viaje.getTrayectoPrincipal().localidad_desde, viaje.getTrayectoPrincipal().localidad_hasta)
    except Exception as e:
        base = 0

    if base == 0:
        try:
            base = getTarifaTrayecto(viaje.categoria_viaje_id, tarifario, viaje.getTrayectoPrincipal().localidad_hasta, viaje.getTrayectoPrincipal().localidad_desde)
        except Exception as e:
            base = 0
    if base is None:
        base = 0
    return base

def getTarifaTrayecto(categoria_viaje, tarifario, desde, hasta):
    if categoria_viaje == 1:
        return TarifaTrayecto.objects.get(tarifario=tarifario, localidad_desde=desde, localidad_hasta=hasta).cat1
    elif categoria_viaje == 2:
        return TarifaTrayecto.objects.get(tarifario=tarifario, localidad_desde=desde, localidad_hasta=hasta).cat2
    elif categoria_viaje == 3:
        return TarifaTrayecto.objects.get(tarifario=tarifario, localidad_desde=desde, localidad_hasta=hasta).cat3
    elif categoria_viaje == 4:
        return TarifaTrayecto.objects.get(tarifario=tarifario, localidad_desde=desde, localidad_hasta=hasta).cat4
    elif categoria_viaje == 5:
        return TarifaTrayecto.objects.get(tarifario=tarifario, localidad_desde=desde, localidad_hasta=hasta).cat5
    elif categoria_viaje == 6:
        return TarifaTrayecto.objects.get(tarifario=tarifario, localidad_desde=desde, localidad_hasta=hasta).cat6
    elif categoria_viaje == 7:
        return TarifaTrayecto.objects.get(tarifario=tarifario, localidad_desde=desde, localidad_hasta=hasta).cat7
    elif categoria_viaje == 8:
        return TarifaTrayecto.objects.get(tarifario=tarifario, localidad_desde=desde, localidad_hasta=hasta).cat8
    elif categoria_viaje == 9:
        return TarifaTrayecto.objects.get(tarifario=tarifario, localidad_desde=desde, localidad_hasta=hasta).cat9

def getTarifaTrayectoExtra(categoria_viaje, tarifario, extra_descripcion):
    if categoria_viaje == 1:
        return TarifaExtra.objects.get(tarifario=tarifario, extra_descripcion=extra_descripcion).cat1
    elif categoria_viaje == 2:
        return TarifaExtra.objects.get(tarifario=tarifario, extra_descripcion=extra_descripcion).cat2
    elif categoria_viaje == 3:
        return TarifaExtra.objects.get(tarifario=tarifario, extra_descripcion=extra_descripcion).cat3
    elif categoria_viaje == 4:
        return TarifaExtra.objects.get(tarifario=tarifario, extra_descripcion=extra_descripcion).cat4
    elif categoria_viaje == 5:
        return TarifaExtra.objects.get(tarifario=tarifario, extra_descripcion=extra_descripcion).cat5
    elif categoria_viaje == 6:
        return TarifaExtra.objects.get(tarifario=tarifario, extra_descripcion=extra_descripcion).cat6
    elif categoria_viaje == 7:
        return TarifaExtra.objects.get(tarifario=tarifario, extra_descripcion=extra_descripcion).cat7
    elif categoria_viaje == 8:
        return TarifaExtra.objects.get(tarifario=tarifario, extra_descripcion=extra_descripcion).cat8
    elif categoria_viaje == 9:
        return TarifaExtra.objects.get(tarifario=tarifario, extra_descripcion=extra_descripcion).cat9


def guardaItemViajeCostoProveedor(monto, tipo_item_viaje, cant, viaje, manual):

    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    base = validateGetTarifaTrayecto('unidad', viaje)
    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje, tipo_items_viaje=tipo_item_viaje)
        monto = 0 if monto == '' else monto
    except ItemViaje.DoesNotExist:
        monto = 0 if monto == '' else monto
        item_viaje_otros = ItemViaje()

    print monto
    print base
    print manual
    monto       = round(float(monto), 2)
    monto_s_iva = monto if manual else round(float(base), 2)
    monto_iva   = round(monto_s_iva * tipo_item_viaje.iva_pct, 2) if manual else round(monto_s_iva * tipo_item_viaje.iva_pct, 2)

    item_viaje_otros.cant               = cant
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

def guardaItemViajeCostoCliente(monto, tipo_item_viaje, cant, viaje, manual):
    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    base = validateGetTarifaTrayecto('centro_costo', viaje)

    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje, tipo_items_viaje=tipo_item_viaje)
        monto = 0 if monto == '' else monto
    except ItemViaje.DoesNotExist:
        monto = 0 if monto == '' else monto
        item_viaje_otros = ItemViaje()

    monto       = round(float(monto), 2)
    monto_s_iva = monto if manual else round(float(base), 2)
    monto_iva   = round(monto_s_iva * tipo_item_viaje.iva_pct, 2) if manual else round(monto_s_iva * tipo_item_viaje.iva_pct, 2)

    item_viaje_otros.cant               = cant
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

def guardaItemViajeMaletas(monto, tipo_item_viaje, cant, viaje, manual):
    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    centro_costo = viaje.centro_costo

    try:
        base = getTarifaTrayectoExtra(viaje.categoria_viaje_id, centro_costo.tarifario, 'maletas')
    except Exception as e:
        base = 0
    if base == None:
        base = 0

    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje,tipo_items_viaje=tipo_item_viaje)
        monto = 0 if monto == '' else monto
    except ItemViaje.DoesNotExist:
        monto = 0 if monto == '' else monto
        item_viaje_otros = ItemViaje()

    monto = round(float(monto), 2)
    monto_s_iva = monto if manual else round(int(cant) * float(base), 2)
    monto_iva   = round(monto_s_iva * tipo_item_viaje.iva_pct, 2) if manual else round(monto_s_iva * tipo_item_viaje.iva_pct, 2)

    item_viaje_otros.cant               = cant
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

def guardaItemViajeMaletasAdmin(monto, tipo_item_viaje, cant, viaje, manual):
    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    unidad = viaje.unidad

    try:
        base = getTarifaTrayectoExtra(viaje.categoria_viaje_id, unidad.tarifario, 'maletas')
    except Exception as e:
        base = 0
    if base == None:
        base = 0
    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje,tipo_items_viaje=tipo_item_viaje)
        monto = 0 if monto == '' else monto
    except ItemViaje.DoesNotExist:
        monto = 0 if monto == '' else monto
        item_viaje_otros = ItemViaje()

    monto = round(float(monto), 2)
    monto_s_iva = monto if manual else round(int(cant) * float(base), 2)
    monto_iva   = round(monto_s_iva * tipo_item_viaje.iva_pct, 2) if manual else round(monto_s_iva * tipo_item_viaje.iva_pct, 2)

    item_viaje_otros.cant               = cant
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

def guardaItemViajeBilingue(monto, tipo_item_viaje, checkbox, viaje, manual, espera, dispo):
    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    base = validateGetTarifaTrayecto('centro_costo', viaje)
    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje, tipo_items_viaje=tipo_item_viaje)
        monto = 0 if monto == '' else monto
        checkbox = 1 if checkbox == 'on' else 0
    except ItemViaje.DoesNotExist:
        if checkbox == '' and monto == '':
            return
        else:
            checkbox = 1 if checkbox == 'on' else 0
            monto = 0 if monto == '' else monto
        item_viaje_otros = ItemViaje()

    monto       = round(float(monto), 2)
    if dispo is None:
        dispo = 0
    try:
        monto_s_iva = monto if manual else round(float(base + espera + dispo) * 0.2, 2)
    except Exception as e:
        monto_s_iva = 0
    monto_iva   = round(monto_s_iva * tipo_item_viaje.iva_pct, 2) if manual else round(monto_s_iva * tipo_item_viaje.iva_pct, 2)

    item_viaje_otros.cant               = checkbox
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

def guardaItemViajeBilingueAdmin(monto, tipo_item_viaje, checkbox, viaje, manual, espera, dispo):
    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    base = validateGetTarifaTrayecto('unidad', viaje)
    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje, tipo_items_viaje=tipo_item_viaje)
        monto = 0 if monto == '' else monto
        checkbox = 1 if checkbox == 'on' else 0
    except ItemViaje.DoesNotExist:
        if checkbox == '' and monto == '':
            return
        else:
            checkbox = 1 if checkbox == 'on' else 0
            monto = 0 if monto == '' else monto
        item_viaje_otros = ItemViaje()


    monto       = round(float(monto), 2)
    if dispo is None:
        dispo = 0
    try:
        monto_s_iva = monto if manual else round(float(base + espera + dispo) * 0.2, 2)
    except Exception as e:
        monto_s_iva = 0
    monto_iva   = round(monto_s_iva * tipo_item_viaje.iva_pct, 2) if manual else round(monto_s_iva * tipo_item_viaje.iva_pct, 2)

    item_viaje_otros.cant               = checkbox
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

def guardaItemViajeEspera(monto, tipo_item_viaje, tiempo, viaje, manual):

    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    centro_costo    = viaje.centro_costo
    try:
        base = getTarifaTrayectoExtra(viaje.categoria_viaje_id, centro_costo.tarifario, 'espera')
    except Exception as e:
        base = 0
    if base == None:
        base = 0

    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje,tipo_items_viaje=tipo_item_viaje)
        monto = '0' if monto == '' else monto
        tiempo = 0 if tiempo == '' else tiempo
    except ItemViaje.DoesNotExist:
        if tiempo == '' and monto == '':
            return
        else:
            monto = '0' if monto == '' else monto
            tiempo = 0 if tiempo == '' else tiempo
        item_viaje_otros = ItemViaje()

    monto       = round(float(monto.replace(',','.')), 2)
    monto_s_iva = monto if manual else round((int(tiempo)/15) * float(base), 2)
    monto_iva   = round(monto_s_iva * tipo_item_viaje.iva_pct, 2) if manual else round(monto_s_iva * tipo_item_viaje.iva_pct, 2)

    item_viaje_otros.cant               = tiempo
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

    return monto_s_iva

def guardaItemViajeEsperaAdmin(monto, tipo_item_viaje, tiempo, viaje, manual):
    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    unidad          = viaje.unidad

    try:
        base = getTarifaTrayectoExtra(viaje.categoria_viaje_id, unidad.tarifario, 'espera')
    except Exception as e:
        base = 0
    if base == None:
        base = 0
    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje,tipo_items_viaje=tipo_item_viaje)
        monto = '0' if monto == '' else monto
        tiempo = 0 if tiempo == '' else tiempo
    except ItemViaje.DoesNotExist:
        if tiempo == '' and monto == '':
            return
        else:
            monto = '0' if monto == '' else monto
            tiempo = 0 if tiempo == '' else tiempo
        item_viaje_otros = ItemViaje()

    monto       = round(float(monto.replace(',','.')), 2)
    monto_s_iva = monto if manual else round((int(tiempo)/15) * float(base), 2)
    monto_iva   = round(monto_s_iva * tipo_item_viaje.iva_pct, 2) if manual else round(monto_s_iva * tipo_item_viaje.iva_pct, 2)

    item_viaje_otros.cant               = tiempo
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

    return monto_s_iva

def guardaItemViajeHsDispo(monto, tipo_item_viaje, tiempo, viaje, manual):
    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    centro_costo = viaje.centro_costo
    try:
        base = getTarifaTrayectoExtra(viaje.categoria_viaje_id, centro_costo.tarifario, 'dispo')
    except Exception as e:
        base = 0
    if base == None:
        base = 0
    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje,tipo_items_viaje=tipo_item_viaje)
        monto = 0 if monto == '' else monto
        tiempo = 0 if tiempo == '' else tiempo
    except ItemViaje.DoesNotExist:
        if monto == '' and tiempo == '':
            return
        else:
            monto = 0 if monto == '' else monto
            tiempo = 0 if tiempo == '' else tiempo
        item_viaje_otros = ItemViaje()

    monto       = round(float(monto), 2)
    monto_s_iva = monto if manual else int(tiempo) * round(float(base))
    monto_iva   = round(monto_s_iva * tipo_item_viaje.iva_pct, 2) if manual else round(monto_s_iva * tipo_item_viaje.iva_pct, 2)

    item_viaje_otros.cant               = tiempo
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

    return monto_s_iva

def guardaItemViajeHsDispoAdmin(monto, tipo_item_viaje, tiempo, viaje, manual):
    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    unidad = viaje.unidad
    try:
        base = getTarifaTrayectoExtra(viaje.categoria_viaje_id, unidad.tarifario, 'dispo')
    except Exception as e:
        base = 0

    if base == None:
        base = 0
    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje,tipo_items_viaje=tipo_item_viaje)
        monto = 0 if monto == '' else monto
        tiempo = 0 if tiempo == '' else tiempo
    except ItemViaje.DoesNotExist:
        if monto == '' and tiempo == '':
            return
        else:
            monto = 0 if monto == '' else monto
            tiempo = 0 if tiempo == '' else tiempo
        item_viaje_otros = ItemViaje()
    monto       = round(float(monto), 2)
    monto_s_iva = monto if manual else int(tiempo) * round(float(base), 2)
    monto_iva   = monto_s_iva * tipo_item_viaje.iva_pct if manual else monto_s_iva * tipo_item_viaje.iva_pct

    item_viaje_otros.cant               = tiempo
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto_s_iva
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto_iva
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

    return monto_s_iva

def guardaItemViaje(monto, tipo_item_viaje, cant, viaje, manual):
    tipo_item_viaje = TipoItemViaje.objects.get(id=tipo_item_viaje)
    try:
        item_viaje_otros = ItemViaje.objects.get(viaje=viaje, tipo_items_viaje=tipo_item_viaje)
        monto = 0 if monto == '' else monto
    except ItemViaje.DoesNotExist:
        if monto == '':
            return
        item_viaje_otros = ItemViaje()

    monto                               = round(float(monto), 2)
    item_viaje_otros.cant               = cant
    item_viaje_otros.monto              = monto
    item_viaje_otros.monto_s_iva        = monto
    item_viaje_otros.viaje              = viaje

    item_viaje_otros.monto_iva          = monto * tipo_item_viaje.iva_pct
    item_viaje_otros.tipo_items_viaje   = tipo_item_viaje
    item_viaje_otros.save()

def guardaViajePasajeroPOST(request):
    viaje                        = Viaje.objects.get(id=request.POST.get('viaje', False))
    viaje_pasajero               = ViajePasajero()
    viaje_pasajero.viaje         = viaje
    viaje_pasajero.pasajero      = Persona.objects.get(id=request.POST.get('pasajero', False))
    viaje_pasajero.pasajero_ppal = False
    viaje_pasajero.save()

    context = {'viaje':viaje}
    return render(request, 'sistema/grillaPasajerosViaje.html', context)

def getViajePasajeros(request):
    viaje = Viaje.objects.get(id=request.POST.get('viaje', False))

    dump = serializers.serialize('json', viaje.getPasajeros())
    return HttpResponse(dump, content_type='application/json')

def deleteViajePasajero(request):
    viaje = Viaje.objects.get(id=request.POST.get('viaje', False))
    ViajePasajero.objects.filter(pasajero_id=request.POST.get('pasajero', False), viaje=viaje).delete()

    context = {'viaje':viaje}
    return render(request, 'sistema/grillaPasajerosViaje.html', context)

def deleteAllViajePasajero(request):
    viaje = Viaje.objects.get(id=request.POST.get('viaje', False))
    ViajePasajero.objects.filter(viaje=viaje).delete()

    context = {'viaje':viaje}
    return render(request, 'sistema/grillaPasajerosViaje.html', context)

def guardaViajePasajero(pasajero, principal, viaje):
    try:
        viaje_pasajero = ViajePasajero.objects.get(viaje=viaje, pasajero_ppal=principal)
    except ViajePasajero.DoesNotExist:
        viaje_pasajero = ViajePasajero()

    viaje_pasajero.viaje         = viaje
    viaje_pasajero.pasajero      = pasajero
    viaje_pasajero.pasajero_ppal = principal
    viaje_pasajero.save()

def guardaObsViaje(input_text, viaje, request):

    if len(viaje.observacionviaje_set.all()) > 0:
        observacion = viaje.observacionviaje_set.all()[0].observacion
        ob_viaje = viaje.observacionviaje_set.all()[0]
    else:
        if input_text == '':
            return
        ob_viaje = ObservacionViaje()
        observacion = Observacion()
        observacion.tipo_observacion = TipoObservacion.objects.get(id=17)

    observacion.fecha = fecha()
    observacion.usuario = request.user
    observacion.texto = input_text
    observacion.save()

    ob_viaje.viaje = viaje
    ob_viaje.observacion = observacion
    ob_viaje.save()

@login_required
def guardarViajeAdjunto(request):
    idViaje = request.POST["idViaje"]
    viaje   = Viaje.objects.get(id=idViaje)
    file    = request.FILES['file']
    error   = False

    options = {
        "maxfilesize": 2 * 2 ** 20,  # 2 Mb
        "minfilesize": 1 * 2 ** 10,  # 1 Kb
        "acceptedformats": (
            "image/jpeg",
            "image/png",
            "application/pdf",
            "application/msword",
        )
    }
    # file size
    if file.size > options["maxfilesize"]:
        error = "maxFileSize"
    #if file.size < options["minfilesize"]:
        #error = "minFileSize"
        # allowed file type
    #if file.content_type not in options["acceptedformats"]:
        #error = "acceptFileTypes"

    response_data = {
        "name": file.name,
        "size": file.size,
        "type": file.content_type
    }

    if error:
        data = {
            'error': '1',
            'msg': error
        }

        dump = json.dumps(data)
        return HttpResponse(dump, content_type='application/json')

    adjunto = Adjunto()
    adjunto.upload = request.FILES['file']
    adjunto.fecha = '2018'
    adjunto.save()

    viajeAdjunto = AdjuntoViaje()
    viajeAdjunto.adjunto = adjunto
    viajeAdjunto.viaje_id = idViaje
    viajeAdjunto.save()

    context = {'viaje': viaje}
    return render(request, 'sistema/grillaAdjuntos.html', context)

def deleteViajeAdjunto(request):
    viaje = Viaje.objects.get(id=request.POST.get('viaje', False))
    AdjuntoViaje.objects.filter(adjunto_id=request.POST.get('adjunto_id', False), viaje=viaje).delete()

    context = {'viaje': viaje}
    return render(request, 'sistema/grillaAdjuntos.html', context)

@login_required
def guardarTrayecto(request, principal=None, idViaje=None):

    if principal:
        trayectos = Trayecto.objects.filter(viaje_id=idViaje)
        viaje = Viaje.objects.get(id=idViaje)
        principal = '1'
    else:
        trayectos = Trayecto.objects.filter(viaje_id=request.POST.get('idViaje', False))
        viaje = Viaje.objects.get(id=request.POST.get('idViaje', False))
        principal = request.POST.get('principal', '')

    if principal != '1' and trayectos.count() == 0:
        data = {
            'error': '1',
            'msg': 'Debes crear un trayecto principal antes de agregar trayectos secundarios.'
        }
        dump = json.dumps(data)
        return HttpResponse(dump, content_type='application/json')
    else:
        if trayectos.count() == 0 and principal == '1':
            trayecto = Trayecto()
        elif trayectos.count() >= 1 and principal == '1':
            trayecto = viaje.getTrayectoPrincipal()
        else:
            id = request.POST.get('id', '')
            if id == '0':
                trayecto = Trayecto()
            else:
				trayecto = Trayecto.objects.get(id=id)


        trayecto.viaje = viaje
        #trayecto.destino_desde = TrayectoDestino.objects.get(id=request.POST.get('desde_destino', False))

        #if request.POST.get('desde_destino', '') != '':
            #trayecto.destino_desde = TrayectoDestino.objects.get(id=request.POST.get('desde_destino', False))

        if request.POST.get('desde_localidad', '') != '':
            trayecto.localidad_desde = Localidad.objects.get(id=request.POST.get('desde_localidad', False))
        if request.POST.get('desde_destino', '') != '':
            trayecto.provincia_desde = Provincia.objects.get(id=request.POST.get('desde_destino', False))
        trayecto.altura_desde = request.POST.get('desde_altura', '')
        trayecto.calle_desde = request.POST.get('desde_calle', '')
        trayecto.entre_desde = request.POST.get('desde_entre', '')
        trayecto.compania_desde = request.POST.get('desde_compania', '')
        trayecto.vuelo_desde = request.POST.get('desde_vuelo', '')

        #trayecto.destino_hasta = TrayectoDestino.objects.get(id=request.POST.get('hasta_destino', False))
        #if request.POST.get('hasta_destino', '') != '':
            #trayecto.destino_hasta = TrayectoDestino.objects.get(id=request.POST.get('hasta_destino', False))

        if request.POST.get('hasta_localidad', '') != '':
            trayecto.localidad_hasta = Localidad.objects.get(id=request.POST.get('hasta_localidad', False))
        if request.POST.get('hasta_destino', '') != '':
            trayecto.provincia_hasta = Provincia.objects.get(id=request.POST.get('hasta_destino', False))
        trayecto.altura_hasta = request.POST.get('hasta_altura', '')
        trayecto.calle_hasta = request.POST.get('hasta_calle', '')
        trayecto.entre_hasta = request.POST.get('hasta_entre', '')
        trayecto.compania_hasta = request.POST.get('hasta_compania', '')
        trayecto.vuelo_hasta = request.POST.get('hasta_vuelo', '')
        if request.POST.get('pasajero', False):
            pasajero = Persona.objects.get(id=request.POST.get('pasajero', False))
            trayecto.pasajero = pasajero
        trayecto.save()

        guardarHistorial(viaje, 'desde_destino', request.POST.get('desde_destino', ''), request.user)
        guardarHistorial(viaje, 'desde_localidad', request.POST.get('desde_localidad', ''), request.user)
        guardarHistorial(viaje, 'desde_provincia', request.POST.get('desde_provincia', ''), request.user)
        guardarHistorial(viaje, 'desde_altura', request.POST.get('desde_altura', ''), request.user)
        guardarHistorial(viaje, 'desde_calle', request.POST.get('desde_calle', ''), request.user)
        guardarHistorial(viaje, 'desde_compania', request.POST.get('desde_compania', ''), request.user)
        guardarHistorial(viaje, 'desde_vuelo', request.POST.get('desde_vuelo', ''), request.user)

        guardarHistorial(viaje, 'hasta_destino', request.POST.get('hasta_destino', ''), request.user)
        guardarHistorial(viaje, 'hasta_localidad', request.POST.get('hasta_localidad', ''), request.user)
        guardarHistorial(viaje, 'hasta_provincia', request.POST.get('hasta_provincia', ''), request.user)
        guardarHistorial(viaje, 'hasta_altura', request.POST.get('hasta_altura', ''), request.user)
        guardarHistorial(viaje, 'hasta_calle', request.POST.get('hasta_calle', ''), request.user)
        guardarHistorial(viaje, 'hasta_compania', request.POST.get('hasta_compania', ''), request.user)
        guardarHistorial(viaje, 'hasta_vuelo', request.POST.get('hasta_vuelo', ''), request.user)



        if principal == '1':
            data = {
                'error': '0',
                'msg': 'Los datos han sido guardados correctamente.'
            }
            #dump = json.dumps(data)
            #return HttpResponse(dump, content_type='application/json')
        else:
            mensaje = ''
            context = {'mensaje': mensaje, 'trayectos': trayectos}
            return render(request, 'sistema/grillaTramos.html', context)

@login_required
def borrarTrayecto(request):
    id = request.POST.get('id', '')
    trayecto = Trayecto.objects.get(id=id)
    trayecto.delete()

    trayectos = Trayecto.objects.filter(viaje_id=request.POST.get('idViaje', False))
    mensaje = ''
    context = {'mensaje': mensaje, 'trayectos': trayectos}
    return render(request, 'sistema/grillaTramos.html', context)

@login_required
def guardarSolicitanteDesdeViaje(request):
    mensaje = ""
    idClienteEnSol = request.POST.get('idClienteEnSol', "")
    cliente = Cliente.objects.get(id=idClienteEnSol)
    idSol = request.POST.get('idSolicitante', "")
    if idSol == "0":
        persona = Persona()
        persona.tipo_persona = TipoPersona.objects.get(id=1)
    else:
        persona = Persona.objects.get(id=idSol)

    persona.nombre = request.POST.get('nombrePasajeroCliente', "")
    persona.apellido = request.POST.get('apellidoPasajeroCliente', "")
    persona.puesto = request.POST.get('puestoSol', "")
    persona.mail = request.POST.get('mailSol', "")
    persona.save()
    telefono = request.POST.get('telefonoSol', "")

    if idSol == "0":
        perCli = PersonaCliente()
        perCli.persona = persona
        perCli.cliente = cliente
        perCli.save()

    if telefono != "" and telefono != "Sin telefono":
        if len(persona.telefonopersona_set.all()) > 0:
            tel = persona.telefonopersona_set.all()[0].telefono
            telcli = persona.telefonopersona_set.all()[0]
        else:
            telcli = TelefonoPersona()
            tel = Telefono()
            tel.tipo_telefono = TipoTelefono.objects.get(tipo_telefono="Principal")

        tel.numero = telefono
        tel.save()

        telcli.persona = persona
        telcli.telefono = tel
        telcli.save()

    context = {'mensaje': mensaje, 'cliente':cliente}
    return render(request, 'sistema/selectSolicitante.html', context)

@login_required
def guardarPasajeroDesdeViaje(request):
    mensaje = ""
    idClientePasajero = request.POST.get('idClientePasajeroModal', "")
    #cliente = Cliente.objects.get(id=idClientePasajero)
    idPasajero = request.POST.get('idPasajeroModal', "")
    if idPasajero == "0":
        persona = Persona()
        persona.tipo_persona = TipoPersona.objects.get(id=2)
    else:
        persona = Persona.objects.get(id=idPasajero)

    persona.nombre = request.POST.get('nombrePasClienteModal', "")
    persona.apellido = request.POST.get('apellidoPasClienteModal', "")
    persona.documento = request.POST.get('documentoPasajeroClienteModal', "")
    persona.mail = request.POST.get('mailPasajeroClienteModal', "")
    persona.nacionalidad = request.POST.get('nacionalidadPasajeroClienteModal', "")
    persona.calle = request.POST.get('callePasajeroClienteModal', "")
    persona.telefono = request.POST.get('telefonoPasajeroClienteModal', "")
    if request.POST.get('pasajeroFrecuente', '') == 'on':
        persona.pasajero_frecuente = True
    else:
        persona.pasajero_frecuente = False
    #persona.altura = request.POST.get('alturaPasajeroCliente', "")
    #persona.piso = request.POST.get('pisoPasajeroCliente', "")
    #persona.cp = request.POST.get('cpPasajeroCliente', "")
    persona.save()


    telefono = persona.telefono
    comentario = request.POST.get('comentarioPasajeroClienteModal', "")

    if idPasajero == "0":
        perCli = PersonaCliente()
        perCli.persona = persona
        perCli.cliente_id = idClientePasajero
        perCli.save()

    if telefono != "" and telefono != "Sin telefono":
        if len(persona.telefonopersona_set.all()) > 0:
            tel = persona.telefonopersona_set.all()[0].telefono
            telcli = persona.telefonopersona_set.all()[0]
        else:
            telcli = TelefonoPersona()
            tel = Telefono()
            tel.tipo_telefono = TipoTelefono.objects.get(tipo_telefono="Principal")

        tel.numero = telefono
        tel.save()

        telcli.persona = persona
        telcli.telefono = tel
        telcli.save()

    if comentario != "":
        if len(persona.observacionpersona_set.all()) > 0:
            obs = persona.observacionpersona_set.all()[0].observacion
            obsper = persona.observacionpersona_set.all()[0]
        else:
            obsper = ObservacionPersona()
            obs = Observacion()
            obs.tipo_observacion = TipoObservacion.objects.get(id=16)

        obs.fecha = fecha()
        obs.usuario = request.user
        obs.texto = comentario
        obs.save()

        obsper.persona = persona
        obsper.observacion = obs
        obsper.save()

    data = {
        'pasajero': persona.id,
        'pasajero_nombre': persona.nombre,
        'pasajero_apellido': persona.apellido,
        'pasajero_telefono': persona.telefono,
        'pasajero_frecuente': persona.pasajero_frecuente
    }
    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')



@login_required
def getTelefonoPasajeroById(request):
    persona = Persona.objects.get(id=request.POST.get('pasajero_id', False))
    data = {
        'telefono': persona.telefono
    }

    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')

@login_required
def altaPersona(request):
	mensaje = ""
	context = {'mensaje': mensaje}
	return render(request, 'sistema/altaPersona.html', context)

@login_required
def editaPersona(request):
	mensaje = ""

	context = {'mensaje': mensaje}
	return render(request, 'sistema/altaPersona.html', context)

@login_required
def guardarOwnerProspect(request):
	persona = Persona()
	persona.nombre = request.POST.get('nombreDuenio', "")
	persona.apellido = request.POST.get('apellidoDuenio', "")
	if request.POST.get('fecNacDuenio', "") != "":
		persona.fecha_nacimiento = getAAAAMMDD(request.POST.get('fecNacDuenio', ""))
	persona.documento = request.POST.get('dniDuenio', "")
	persona.mail = request.POST.get('mailDuenio', "")
	persona.calle = request.POST.get('domDuenio', "")
	persona.tipo_persona_id = 4
	persona.save()

	telefono = request.POST.get('telefonoDuenio', "")
	if telefono != "":
		tel = Telefono()
		tel.tipo_telefono = TipoTelefono.objects.get(tipo_telefono="Principal")
		tel.numero = telefono
		tel.save()

		telcli = TelefonoPersona()
		telcli.persona = persona
		telcli.telefono = tel
		telcli.save()

	data = {
		'persona_id': persona.id,
		'persona_nombre': persona.nombre,
		'persona_apellido':persona.apellido
	}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

def guardarChoferProspect(request):
	persona = Persona()
	persona.nombre = request.POST.get('nombreChofer', "")
	persona.apellido = request.POST.get('apellidoChofer', "")
	persona.porcentaje_viaje = request.POST.get('porcentajeChofer', "")
	if request.POST.get('fecNacChofer', "") != "":
		persona.fecha_nacimiento = getAAAAMMDD(request.POST.get('fecNacChofer', ""))
	persona.documento = request.POST.get('dniChofer', "")
	persona.mail = request.POST.get('mailChofer', "")
	persona.calle = request.POST.get('domChofer', "")
	persona.tipo_persona_id = 3
	persona.save()

	telefono = request.POST.get('telefonoChofer', "")
	if telefono != "":
		tel = Telefono()
		tel.tipo_telefono = TipoTelefono.objects.get(tipo_telefono="Principal")
		tel.numero = telefono
		tel.save()

		telcli = TelefonoPersona()
		telcli.persona = persona
		telcli.telefono = tel
		telcli.save()

	data = {
		'persona_id': persona.id,
		'persona_nombre': persona.nombre,
		'persona_apellido':persona.apellido,
		'persona_porcentaje':persona.porcentaje_viaje
	}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

@login_required
def guardarCentroCostoProspectDesdeViaje(request):
    idCC = request.POST.get('idClienteCC', "")
    if idCC == "0":
        cc = CentroCosto()
        cliente = Cliente.objects.get(id=request.POST.get('idClienteEnCC', ""))
        cc.cliente = cliente
    else:
        cc = CentroCosto.objects.get(id=idCC)
        cliente = cc.cliente

    cc.nombre = request.POST.get('codigoCCCliente', "")
    cc.fecha_inicio = getAAAAMMDD(request.POST.get('desdeCC', ""))
    cc.fecha_fin = getAAAAMMDD(request.POST.get('hastaCC', ""))
    cc.descripcion = request.POST.get('descripcionCCCliente', "")
    cc.tarifario = Tarifario.objects.get(id=request.POST.get('selectTarifariosCCCliente', ""))
    cc.save()
    idCliente = request.POST.get('idClienteEnCC', False)
    centrosDeCosto = []
    print idCliente
    if idCliente:
        centrosDeCosto = CentroCosto.objects.filter(cliente_id=idCliente)

    context = {'centrosDeCosto': centrosDeCosto}
    return render(request, 'sistema/selectCentroCosto.html', context)

@login_required
def guardarCentroCostoProspect(request):
	mensaje = ""
	trayectos = ""
	idCC = request.POST.get('idClienteCC', "")
	if idCC == "0":
		cc = CentroCosto()
		cliente = Cliente.objects.get(id=request.POST.get('idClienteEnCC', ""))
		cc.cliente = cliente
	else:
		cc = CentroCosto.objects.get(id=idCC)
		cliente = cc.cliente

	cc.nombre = request.POST.get('codigoCCCliente', "")
	cc.fecha_inicio = getAAAAMMDD(request.POST.get('desdeCC', ""))
	cc.fecha_fin = getAAAAMMDD(request.POST.get('hastaCC', ""))
	cc.descripcion = request.POST.get('descripcionCCCliente', "")
	cc.tarifario = Tarifario.objects.get(id=request.POST.get('selectTarifariosCCCliente', ""))
	cc.save()

	context = {'mensaje': mensaje, 'trayectos': trayectos, 'cliente':cliente}
	return render(request, 'sistema/grillaCentroCostos.html', context)

@login_required
def checkLocalidadFlag(request):
    if request.POST.get('localidad_id', "") != '':
        localidad = Localidad.objects.get(id=request.POST.get('localidad_id', ""))
        data = {
            'terminal_flag': localidad.terminal_flag
        }
    else:
        data = {
            'terminal_flag': 0
        }
    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')


@login_required
def guardarSolicitanteProspect(request):
	mensaje = ""
	idClienteEnSol = request.POST.get('idClienteEnSol', "")
	cliente = Cliente.objects.get(id=idClienteEnSol)
	idSol = request.POST.get('idSolicitante', "")
	if idSol == "0":
		persona = Persona()
		persona.tipo_persona = TipoPersona.objects.get(id=1)
	else:
		persona = Persona.objects.get(id=idSol)

	persona.nombre = request.POST.get('nombrePasajeroCliente', "")
	persona.apellido = request.POST.get('apellidoPasajeroCliente', "")
	persona.puesto = request.POST.get('puestoSol', "")
	persona.mail = request.POST.get('mailSol', "")
	persona.save()
	telefono = request.POST.get('telefonoSol', "")

	if idSol == "0":
		perCli = PersonaCliente()
		perCli.persona = persona
		perCli.cliente = cliente
		perCli.save()

	if telefono != "" and telefono != "Sin telefono":
		if len(persona.telefonopersona_set.all()) > 0:
			tel = persona.telefonopersona_set.all()[0].telefono
			telcli = persona.telefonopersona_set.all()[0]
		else:
			telcli = TelefonoPersona()
			tel = Telefono()
			tel.tipo_telefono = TipoTelefono.objects.get(tipo_telefono="Principal")

		tel.numero = telefono
		tel.save()

		telcli.persona = persona
		telcli.telefono = tel
		telcli.save()

	context = {'mensaje': mensaje, 'cliente':cliente}
	return render(request, 'sistema/grillaSolicitantes.html', context)

@login_required
def guardarPasajeroProspect(request):
	mensaje = ""
	idClientePasajero = request.POST.get('idClientePasajero', "")

	cliente = Cliente.objects.get(id=idClientePasajero)
	idPasajero = request.POST.get('idPasajero', "")
	if idPasajero == "0":
		persona = Persona()
		persona.tipo_persona = TipoPersona.objects.get(id=2)
	else:
		persona = Persona.objects.get(id=idPasajero)

	persona.nombre = request.POST.get('nombrePasCliente', "")
	persona.apellido = request.POST.get('apellidoPasCliente', "")
	persona.documento = request.POST.get('documentoPasajeroCliente', "")
	persona.mail = request.POST.get('mailPasajeroCliente', "")
	persona.nacionalidad = request.POST.get('nacionalidadPasajeroCliente', "")
	persona.direccion = request.POST.get('callePasajeroCliente', "")
	persona.telefono = request.POST.get('telefonoPasajeroCliente', "")
	if request.POST.get('pasajeroFrecuente', '') == 'on':
		persona.pasajero_frecuente = True
	else:
		persona.pasajero_frecuente = False
	persona.save()
	telefono = persona.telefono
	comentario = request.POST.get('comentarioPasajeroCliente', "")

	if idPasajero == "0":
		perCli = PersonaCliente()
		perCli.persona = persona
		perCli.cliente = cliente
		perCli.save()

	if telefono != "" and telefono != "Sin telefono":
		if len(persona.telefonopersona_set.all()) > 0:
			tel = persona.telefonopersona_set.all()[0].telefono
			telcli = persona.telefonopersona_set.all()[0]
		else:
			telcli = TelefonoPersona()
			tel = Telefono()
			tel.tipo_telefono = TipoTelefono.objects.get(tipo_telefono="Principal")

		tel.numero = telefono
		tel.save()

		telcli.persona = persona
		telcli.telefono = tel
		telcli.save()

	if comentario != "":
		if len(persona.observacionpersona_set.all()) > 0:
			obs = persona.observacionpersona_set.all()[0].observacion
			obsper = persona.observacionpersona_set.all()[0]
		else:
			obsper = ObservacionPersona()
			obs = Observacion()
			obs.tipo_observacion = TipoObservacion.objects.get(id=16)

		obs.fecha = fecha()
		obs.usuario = request.user
		obs.texto = comentario
		obs.save()

		obsper.persona = persona
		obsper.observacion = obs
		obsper.save()

	cliente_id = cliente.id
	pasajeros = cliente.getPasajeros()

	context = {'mensaje': mensaje, 'cliente_id':cliente_id, 'pasajeros':pasajeros}
	return render(request, 'sistema/grillaPasajeros.html', context)

@login_required
def listadoCliente(request, **kwargs):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	clientes = Cliente.objects.filter(baja=False)
	context = {'clientes': clientes}
	return render(request, 'sistema/listadoCliente.html', context)

@login_required
def cliente(request):
	mensaje = ""
	idCliente = request.GET.get('idCliente', "")
	cliente = Cliente.objects.get(id=idCliente)
	tarifarios = Tarifario.objects.all()
	ivas = Iva.objects.all()
	condiciones = CondicionPago.objects.all()
	centrosDeCosto = cliente.getCentroCostos()
	context = {'mensaje': mensaje, 'cliente':cliente, 'tarifarios':tarifarios,"ivas":ivas, "condiciones": condiciones, 'centrosDeCosto':centrosDeCosto }
	return render(request, 'sistema/cliente.html', context)

@login_required
def altaCliente(request):
	mensaje = ""
	cliente = Cliente()
	cliente.id = 0
	tarifarios = Tarifario.objects.all()
	context = {'mensaje': mensaje, 'tarifarios':tarifarios, 'cliente': cliente }
	return render(request, 'sistema/cliente.html', context)

@login_required
def guardarCliente(request):
	idCliente = request.POST.get('idCliente', "")
	if idCliente == "0":
		cliente = Cliente()
		tel = Telefono()
		telcli = TelefonoCliente()
	else:
		cliente = Cliente.objects.get(id=idCliente)
		if len(cliente.telefonocliente_set.all()) > 0:
			tel = cliente.telefonocliente_set.all()[0].telefono
			telcli = cliente.telefonocliente_set.all()[0]
		else:
			tel = Telefono()
			telcli = TelefonoCliente()

	cliente.razon_social = request.POST.get('razonSocial', "")
	cliente.cuil = request.POST.get('cuil', "")
	cliente.calle = request.POST.get('calle', "")
	cliente.altura = request.POST.get('altura',"")
	cliente.piso = request.POST.get('piso', "")
	cliente.depto = request.POST.get('depto', "")
	cliente.cp = request.POST.get('cp', "")
	cliente.localidad = request.POST.get('localidad', "")
	cliente.provincia = request.POST.get('provincia', "")
	iva = request.POST.get('iva', "")
	if iva != "":
		cliente.iva = Iva.objects.get(id=iva)
	cond = request.POST.get('condicionPago', "")
	if cond != "":
		cliente.condicion_pago = CondicionPago.objects.get(id=cond)
	cliente.dias_fechas_facturas = request.POST.get('diasFechaFactura', "")
	cliente.alias = request.POST.get('alias', "")
	cliente.cbu = request.POST.get('cbu', "")
	if request.POST.get('clienteMoroso', "") == 'on':
		cliente.moroso = True
	else:
		cliente.moroso = False
	cliente.save()

	if request.POST.get('telefono', False) != "":
		tel.tipo_telefono = TipoTelefono.objects.get(tipo_telefono="Principal")
		tel.numero = request.POST.get('telefono', False)
		tel.save()
		telcli.cliente = cliente
		telcli.telefono = tel
		telcli.save()

	url = '/sistema/cliente/?idCliente='+str(cliente.id)
	return redirect(url)

@login_required
def editaCliente(request):
	mensaje = ""

	context = {'mensaje': mensaje}
	return render(request, 'sistema/cliente.html', context)

@login_required
def eliminarCliente(request):
	idCliente = request.POST.get('idClienteEliminar', False)
	cliente = Cliente.objects.get(id=idCliente)
	cliente.baja = True
	cliente.save()

	return redirect('listadoCliente')


@login_required
def guardarObservacionViaje(request):
    mensaje = ""

    idViaje = request.POST.get('idViajeObser', False)
    text_obs = request.POST.get('text_obs', False)
    hora_obs = request.POST.get('hora_obs', '')
    detalle_obs = request.POST.get('detalle_obs', False)
    observacion = Observacion()
    observacion.fecha = hora_obs
    observacion.usuario = request.user
    observacion.texto = text_obs
    observacion.tipo_observacion = TipoObservacion.objects.get(id=detalle_obs)
    observacion.save()

    viaje = Viaje.objects.get(id=idViaje)

    obcl = ObservacionViaje()
    obcl.observacion = observacion
    obcl.viaje = viaje
    obcl.save()

    context = {'mensaje': mensaje, 'viaje':viaje}
    return render(request, 'sistema/grillaObservacionesViaje.html', context)


@login_required
def guardarObservacionCliente(request):
	mensaje = ""

	idCliente = request.POST.get('idClienteObser', False)
	observ = request.POST.get('textAreaObservacion', False)
	observacion = Observacion()
	observacion.fecha = fecha()
	observacion.usuario = request.user
	observacion.texto = observ
	observacion.save()

	cliente = Cliente.objects.get(id=idCliente)

	obcl = ObservacionCliente()
	obcl.observacion = observacion
	obcl.cliente = cliente
	obcl.save()

	context = {'mensaje': mensaje, 'objeto':cliente}
	return render(request, 'sistema/grillaObservaciones.html', context)


@login_required
def getHistorial(request):
    mensaje = ""

    historial = ViajeHistorial.objects.filter(viaje_id=request.POST.get('idViaje', False)).order_by('-id')
    context = {'mensaje': mensaje, 'historial':historial}
    return render(request, 'sistema/grillaHistorial.html', context)


@login_required
def getViajesFuturosPorFecha(request):
    mensaje = ""
    date = getAAAAMMDD(request.POST.get('date', False))

    estados_get_seleccionados = request.POST.getlist('estados_selecionados[]', False)
    estados_seleccionados = []
    if estados_get_seleccionados == False:
		estados_seleccionados = [1, 2, 3, 4, 10]
    else:
        for i in estados_get_seleccionados:
			estados_seleccionados.append(i)

	viajes = Viaje.objects.filter(fecha=date).filter(estado__in=estados_seleccionados)
    for v in viajes:
        for p in v.getPasajeros():
            print p
    context = {'mensaje': mensaje, 'viajes':viajes}
    return render(request, 'sistema/grillaViajesFuturos.html', context)

@login_required
def getViajesEnProgresoPorFecha(request):
    mensaje = ""

    date = getAAAAMMDD(request.POST.get('date', False))
    viajes = Viaje.objects.filter(fecha=date)
    context = {'mensaje': mensaje, 'viajes':viajes}
    return render(request, 'sistema/grillaViajesEnProgreso.html', context)

@login_required
def editEstadoViajeAsignaciones(request):
    mensaje = ""
    viajes_get_seleccionado = request.POST.getlist('viajes_seleccionado[]')
    viajes_seleccionados = []
    for i in viajes_get_seleccionado:
        viajes_seleccionados.append(i)
    Viaje.objects.filter(id__in=viajes_seleccionados).update(estado=Estado.objects.get(id=request.POST.get('estado_seleccionado', False)))

    data = {
		'error': '0',
		'msg': mensaje
	}

    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')

@login_required
def editEstadoViaje(request):
    mensaje = ""
    viaje = Viaje.objects.get(id=request.POST.get('viaje_seleccionado', False))
    viaje.estado = Estado.objects.get(id=request.POST.get('estado_seleccionado', False))
    viaje.save()
    data = {
		'error': '0',
		'msg': mensaje
	}

    dump = json.dumps(data)
    return HttpResponse(dump, content_type='application/json')

@login_required
def guardarMailCliente(request):
	mensaje = ""

	idCliente = request.POST.get('idClienteMail', False)
	nombre = request.POST.get('nombreMailCliente', False)
	mail_txt = request.POST.get('mailCliente', False)
	comentario = request.POST.get('comentarioMailCliente', False)
	mail = Mail()
	mail.mail = mail_txt
	mail.nombre = nombre
	mail.comentario = comentario
	mail.save()

	cliente = Cliente.objects.get(id=idCliente)

	mlcl = MailCliente()
	mlcl.mail = mail
	mlcl.cliente = cliente
	mlcl.save()

	context = {'mensaje': mensaje, 'objeto':cliente}
	return render(request, 'sistema/grillaMails.html', context)

@login_required
def guardarObservacionPersona(request):
	mensaje = ""

	idPersona = request.POST.get('idPersonaObser', False)
	observ = request.POST.get('textAreaObservacion', False)
	observacion = Observacion()
	observacion.fecha = fecha()
	observacion.usuario = request.user
	observacion.texto = observ
	observacion.save()

	persona = Persona.objects.get(id=idPersona)

	obpe = ObservacionPersona()
	obpe.observacion = observacion
	obpe.persona = persona
	obpe.save()

	context = {'mensaje': mensaje, 'objeto':persona}
	return render(request, 'sistema/grillaObservaciones.html', context)

@login_required
def listadoProvedor(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	provedores = Persona.objects.filter(tipo_persona__id__in=[1,2],baja=False)
	context = {'provedores': provedores}
	return render(request, 'sistema/listadoProvedor.html', context)

@login_required
def provedor(request):
	mensaje = ""
	idProv = request.GET.get('idProv', "")
	prov = Persona.objects.get(id=idProv)
	estadosCivil = EstadoCivil.objects.all()
	context = {'prov': prov, 'estadosCivil':estadosCivil}
	return render(request, 'sistema/provedor.html', context)

@login_required
def guardarProvedor(request):
	mensaje = ""
	idProv = request.POST.get('idProv', "")
	if idProv == "0":
		persona = Persona()
		persona.tipo_persona_id = request.POST.get('provTipoPersona', "")
	else:
		persona = Persona.objects.get(id=idProv)

	persona.nombre = request.POST.get('provNombre', "")
	persona.apellido = request.POST.get('provApellido', "")
	persona.documento = request.POST.get('provDNI', "")
	if request.POST.get('provNacimiento', "") == "":
		persona.fecha_nacimiento = ""
	else:
		persona.fecha_nacimiento = getAAAAMMDD(request.POST.get('provNacimiento', ""))
	persona.telefono = request.POST.get('provTelefono', "")
	persona.mail = request.POST.get('provMail', "")
	persona.estado_civil_id = request.POST.get('provEstadoCivil', "")
	persona.direccion = request.POST.get('direccion', "")
	persona.cp = request.POST.get('cp', "")
	persona.localidad = request.POST.get('localidad', "")
	persona.provincia = request.POST.get('provincia', "")
	persona.save()

	url = '/sistema/provedor/?idProv='+str(persona.id)
	return redirect(url)

@login_required
def altaProvedor(request):
	prov = Persona()
	prov.id = 0
	estadosCivil = EstadoCivil.objects.all()
	context = {'estadosCivil': estadosCivil, 'prov':prov}
	return render(request, 'sistema/provedor.html', context)

@login_required
def borrarProvedor(request):
	mensaje = ""
	idProv = request.GET.get('idProv', "")
	prov = Persona.objects.get(id=idProv)
	prov.baja = True
	prov.save()
	return redirect('listadoProvedor')

@login_required
def unidad(request):
	mensaje = ""
	idUnidad = request.GET.get('idUnidad', "")
	unidad = Unidad.objects.get(id=idUnidad)
	tipo_licencias = TipoLicencia.objects.all()
	tarifarios = Tarifario.objects.filter(baja=False)
	ids_fake = []
	ids_fake.append(unidad.id_fake)
	unidades = Unidad.objects.filter(baja=False).values_list('id_fake', flat=True)
	unidades = map(lambda x:int(x), unidades)
	for number in range(1100):
		if number not in unidades:
			ids_fake.append(number)
	context = {'mensaje': mensaje, 'unidad': unidad, 'tipo_licencias':tipo_licencias,'ids_fake':ids_fake,'tarifarios':tarifarios}
	return render(request, 'sistema/unidad.html', context)

@login_required
def altaUnidad(request):
	mensaje = ""
	ids_fake = []
	unidad = Unidad()
	unidad.id = 0
	unidades = Unidad.objects.values_list('id_fake', flat=True)
	unidades = map(lambda x:int(x), unidades)
	tarifarios = Tarifario.objects.filter(baja=False)
	for number in range(1100):
		if number not in unidades:
			ids_fake.append(number)

	context = {'mensaje': mensaje, 'unidad': unidad, 'ids_fake':ids_fake, 'tarifarios':tarifarios}
	return render(request, 'sistema/unidad.html', context)

@login_required
def asignarIdFake(request):
	unidades = Unidad.objects.filter(baja=False)
	idFake = 0
	for u in unidades:
		idFake = idFake + 1
		u.id_fake = idFake
		u.save()
	context = {'unidades':unidades}
	return render(request, 'sistema/listadoUnidad.html', context)

@login_required
def listadoUnidad(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	unidades = Unidad.objects.filter(baja=False)
	context = {'unidades':unidades}
	return render(request, 'sistema/listadoUnidad.html', context)

@login_required
def guardarUnidad(request):
	idUnidad = request.POST.get('idUnidad', "")
	if idUnidad == "0":
		unidad = Unidad()
		vehiculo = Vehiculo()
	else:
		unidad = Unidad.objects.get(id=idUnidad)
		if unidad.vehiculo:
			vehiculo = unidad.vehiculo
		else:
			vehiculo = Vehiculo()


	unidad.id_fake = request.POST.get('selectIdFake', "")
	unidad.identificacion = request.POST.get('identificacion', "")
	unidad.alias = request.POST.get('alias', "")
	unidad.mail = request.POST.get('mail', "")
	unidad.telefono = request.POST.get('telefono', "")
	unidad.documento = request.POST.get('documento', "")
	if request.POST.get('fecha_nac', "") != "":
		unidad.fecha_nacimiento = getAAAAMMDD(request.POST.get('fecha_nac', ""))

	unidad.calle = request.POST.get('calle', "")

	if request.POST.get('selectTarifario', "") != "":
		unidad.tarifario = Tarifario.objects.get(id=request.POST.get('selectTarifario', ""))

	if request.POST.get('unidadPropia', "") == "on":
		unidad.unidad_propia = True
	else:
		unidad.unidad_propia = False

	if request.POST.get('patente', "") != "":
		vehiculo.patente = request.POST.get('patente', "")
		vehiculo.modelo = request.POST.get('modelo', "")
		vehiculo.marca = request.POST.get('marca', "")
		vehiculo.color = request.POST.get('color', "")
		vehiculo.ano = request.POST.get('year', "")
		vehiculo.puertas = request.POST.get('puertas', "")
		vehiculo.nro_motor = request.POST.get('nro_motor', "")
		vehiculo.nro_chasis = request.POST.get('nro_chasis', "")
		vehiculo.save()
		unidad.vehiculo = vehiculo
	unidad.save()
	url = '/sistema/unidad/?idUnidad='+str(unidad.id)
	return redirect(url)

@login_required
def eliminarUnidad(request):
	idUnidad = request.POST.get('idUnidadEliminar', False)
	unidad = Unidad.objects.get(id=idUnidad)
	unidad.baja = True
	unidad.save()
	return redirect('listadoUnidad')

@login_required
def guardarObservacionUnidad(request):
	mensaje = ""

	idUnidad = request.POST.get('idUnidadObser', False)
	observ = request.POST.get('textAreaObservacion', False)
	observacion = Observacion()
	observacion.fecha = fecha()
	observacion.usuario = request.user
	observacion.texto = observ
	observacion.save()

	unidad = Unidad.objects.get(id=idUnidad)

	obcl = ObservacionUnidad()
	obcl.observacion = observacion
	obcl.unidad = unidad
	obcl.save()

	context = {'mensaje': mensaje, 'objeto':unidad}
	return render(request, 'sistema/grillaObservaciones.html', context)

@login_required
def guardarLicenciaUnidad(request):
	idLicencia = request.POST.get('idLicencia', False)
	descripcion = request.POST.get('descripcionLicencia', False)
	tipo = TipoLicencia.objects.get(id=request.POST.get('tipoLicencia', False))
	fv = request.POST.get('vencimientoLicencia', False)
	fecha = fv[6:10] + fv[3:5] + fv[0:2]

	if idLicencia == "0":
		licencia = Licencia()
	else:
		licencia = Licencia.objects.get(id=idLicencia)

	unidad = Unidad.objects.get(id=request.POST.get('idUnidad', False))

	licencia.comentario = descripcion
	licencia.tipo_licencia = tipo
	licencia.fecha_vencimiento = fecha
	licencia.unidad = unidad
	licencia.save()

	context = {'unidad':unidad}
	return render(request, 'sistema/grillaLicencias.html', context)

@login_required
def eliminarLicencia(request):
	idLicencia = request.GET.get('idLicencia', False)
	Licencia.objects.get(id=idLicencia).delete()
	return redirect('listadoLicencia')

@login_required
def eliminarLicenciaPropect(request):
	idLicencia = request.POST.get('idLicencia', False)
	idUnidad = request.POST.get('idUnidad', False)
	Licencia.objects.get(id=idLicencia).delete()
	unidad = Unidad.objects.get(id=idUnidad)
	context = {'unidad':unidad}
	return render(request, 'sistema/grillaLicencias.html', context)

@login_required
def contacto(request):
	mensaje = ""

	context = {'mensaje': mensaje}
	return render(request, 'sistema/contacto.html', context)

@login_required
def editaContacto(request):
	mensaje = ""

	context = {'mensaje': mensaje}
	return render(request, 'sistema/contacto.html', context)

@login_required
def listadoContacto(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	contactos = Persona.objects.filter(tipo_persona__in=[1,2], baja=False)
	context = {'contactos': contactos}
	return render(request, 'sistema/listadoContacto.html', context)

@login_required
def altaCentroDeCosto(request):
	clientes = Cliente.objects.filter(baja=False)
	tarifarios = Tarifario.objects.filter(baja=False)
	cc = CentroCosto()
	cc.id = 0
	context = {'clientes': clientes, 'tarifarios':tarifarios, 'cc':cc}
	return render(request, 'sistema/centroDeCosto.html', context)

@login_required
def guardarCC(request):
	idCC = request.POST.get('idCC', False)
	cliente_id = request.POST.get('clientes', False)
	tarifario_id = request.POST.get('tarifarios', False)
	codigo = request.POST.get('codigo', False)
	descripcion = request.POST.get('descripcion', False)
	desdeCC = request.POST.get('desdeCC', False)
	hastaCC = request.POST.get('hastaCC', False)

	if idCC == "0":
		cc = CentroCosto()
		request.session['estadoCC'] = 'nuevo'
	else:
		cc = CentroCosto.objects.get(id=idCC)
		request.session['estadoCC'] = 'editado'

	cc.nombre = codigo
	cc.descripcion = descripcion
	cc.cliente = Cliente.objects.get(id=cliente_id)
	cc.tarifario = Tarifario.objects.get(id=tarifario_id)
	cc.fecha_inicio = getAAAAMMDD(desdeCC)
	cc.fecha_fin = getAAAAMMDD(hastaCC)
	cc.save()

	url = '/sistema/centroCosto/?idCC='+str(cc.id)
	return redirect(url)

@login_required
def centroCosto(request):
	estado = request.session.get('estadoCC', '')
	idCC = request.GET.get('idCC', "")
	cc = CentroCosto.objects.get(id=idCC)
	clientes = Cliente.objects.all()
	tarifarios = Tarifario.objects.all()
	request.session['estadoCC'] = ''

	context = {'clientes':clientes, 'estado':estado, 'cc':cc, 'clientes':clientes, 'tarifarios':tarifarios}
	return render(request, 'sistema/centroDeCosto.html', context)

@login_required
def validarCodigoCentroCosto(request):
	codigoCC = request.GET.get('codigoCC', False)
	cc = CentroCosto.objects.filter(nombre=codigoCC)
	if cc:
		data = {'mensaje': 'El codigo seleccionado ya existe.'}
	else:
		data = {'mensaje': ''}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

@login_required
def editaCentroDeCosto(request):
	mensaje = ""

	context = {'mensaje': mensaje}
	return render(request, 'sistema/centroDeCosto.html', context)

@login_required
def listadoCentroDeCosto(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	centroCostos = CentroCosto.objects.filter(baja=False)

	context = {'centroCostos': centroCostos}
	return render(request, 'sistema/listadoCentroDeCosto.html', context)

@login_required
def buscarCentroDeCostos(request):
	desde = request.POST.get('desde', "")
	hasta = request.POST.get('hasta', "")
	fechaDesde =  getAAAAMMDD(desde)
	fechaHasta =  getAAAAMMDD(hasta)
	viajes = Viaje.objects.filter(fecha__lte=fechaHasta, fecha__gte=fechaDesde,centro_costo__isnull=False)
	centroCostosIds = []
	centroCostos = []
	for v in viajes:
		if v.centro_costo.id not in centroCostosIds:
			centroCostos.append(v.centro_costo)
			centroCostosIds.append(v.centro_costo.id)

	context = {'centroCostos': centroCostos}
	return render(request, 'sistema/grillaListadoCentroDeCosto.html', context)


@login_required
def listadoTarifario(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	mensaje = ""
	tarifarios = Tarifario.objects.filter(baja=False)
	context = {'tarifarios': tarifarios}
	return render(request, 'sistema/listadoTarifario.html', context)

@login_required
def tarifario(request):
	mensaje = ""
	idTarifario = request.GET.get('idTarifario', "")
	tarifario = Tarifario.objects.get(id=idTarifario)
	context = {'tarifario': tarifario}
	return render(request, 'sistema/tarifario.html', context)

@login_required
def guardarTarifario(request):
	mensaje = ""
	idTarifario = request.POST.get('idTarifario', "")
	nombre = request.POST.get('nombre', "")
	tarifario = Tarifario.objects.get(id=idTarifario)
	tarifario.nombre = nombre
	tarifario.save()
	url = '/sistema/tarifario/?idTarifario='+str(tarifario.id)
	return redirect(url)

@login_required
def editarTarifaTrayecto(request):
	idTarifaTrayecto = request.POST.get('idTarifaTrayecto', "")
	localidades = Localidad.objects.all().values_list('id', 'nombre').order_by('nombre')
	localidades = map(lambda localidades:(int(localidades[0]),localidades[1]), localidades)
	if idTarifaTrayecto == "0":
		tramoTarifa = TarifaTrayecto()
		tramoTarifa.id = 0
		idDesde = ""
		idHasta = ""
	else:
		tramoTarifa = TarifaTrayecto.objects.get(id=idTarifaTrayecto)
		idDesde = tramoTarifa.localidad_desde.id
		idHasta = tramoTarifa.localidad_hasta.id
	context = {'tramoTarifa': tramoTarifa, 'localidades':localidades, 'idDesde':idDesde, 'idHasta': idHasta}
	return render(request, 'sistema/tarifaTrayecto.html', context)

@login_required
def validarTarifaTrayecto(request):
	response_data = {}
	idTarifario = request.POST.get('idTarifario', "")
	localidadDesde = request.POST.get('localidadDesde', "")
	localidadHasta = request.POST.get('localidadHasta', "")

	response_data['result'] = 'ok'
	response_data['message'] = ''

	print '*-*-*-*'
	print localidadDesde
	print localidadHasta

	if localidadDesde == "":
		response_data['result'] = 'ok'
		response_data['message'] = 'falta campo desde'
	if localidadHasta == "":
		response_data['result'] = 'ok'
		response_data['message'] = 'falta campo hasta'

	tarifario = Tarifario.objects.get(id=idTarifario)
	for t in tarifario.getTarifaViaje():
		if str(t.localidad_desde.id) == localidadDesde and str(t.localidad_hasta.id) == localidadHasta:
			response_data['result'] = 'error'
			response_data['message'] = 'Ya existe esta combinacion de localidades.'
		if str(t.localidad_desde.id) == localidadHasta and str(t.localidad_hasta.id) == localidadDesde:
			response_data['result'] = 'error'
			response_data['message'] = 'Ya existe esta combinacion de localidades.'

	return JsonResponse(response_data)

@login_required
def editarTarifaExtra(request):
	idTarifaExtra = request.POST.get('idTarifaExtra', "")
	if idTarifaExtra == "0":
		tarifaExtra = TarifaExtra()
		tarifaExtra.id = 0
		tarifaExtra.extra_descripcion = ""
	else:
		tarifaExtra = TarifaExtra.objects.get(id=idTarifaExtra)

	context = {'tarifaExtra': tarifaExtra}
	return render(request, 'sistema/tarifaExtra.html', context)

@login_required
def eliminarTarifaTrayecto(request):
	idTarifaTrayecto = request.POST.get('idTarifaTrayecto', "")
	tarifaTrayecto = TarifaTrayecto.objects.get(id=idTarifaTrayecto)
	idTarifario = tarifaTrayecto.tarifario.id

	tarifaTrayecto.delete()
	tarifario = Tarifario.objects.get(id=idTarifario)
	context = {'tarifario': tarifario}
	return render(request, 'sistema/grillaTrayectoTarifa.html', context)


@login_required
def guardarTarifaTrayecto(request):
	idTarifario		 = request.POST.get('idTarifario', "")
	idTarifaTrayecto = request.POST.get('idTarifaTrayecto', "")

	if idTarifaTrayecto == "0":
		localidadDesde = request.POST.get('localidadDesde', "")
		localidadHasta = request.POST.get('localidadHasta', "")
		tarifaTrayecto = TarifaTrayecto()
		tarifaTrayecto.localidad_desde = Localidad.objects.get(id=localidadDesde)
		tarifaTrayecto.localidad_hasta = Localidad.objects.get(id=localidadHasta)
		tarifaTrayecto.tarifario = Tarifario.objects.get(id=idTarifario)
	else:
		tarifaTrayecto = TarifaTrayecto.objects.get(id=idTarifaTrayecto)

	for x in range(9):
		cat = x+1
		nameTramo='tramo'+str(cat)
		valor = request.POST.get(nameTramo, "0")
		if valor == "":
			valor = 0
		if cat == 1:
			tarifaTrayecto.cat1 = float(valor)
		if cat == 2:
			tarifaTrayecto.cat2 = float(valor)
		if cat == 3:
			tarifaTrayecto.cat3 = float(valor)
		if cat == 4:
			tarifaTrayecto.cat4 = float(valor)
		if cat == 5:
			tarifaTrayecto.cat5 = float(valor)
		if cat == 6:
			tarifaTrayecto.cat6 = float(valor)
		if cat == 7:
			tarifaTrayecto.cat7 = float(valor)
		if cat == 8:
			tarifaTrayecto.cat8 = float(valor)

	tarifaTrayecto.save()

	tarifario = Tarifario.objects.get(id=idTarifario)
	context = {'tarifario': tarifario}
	return render(request, 'sistema/grillaTrayectoTarifa.html', context)

@login_required
def guardarTarifaExtra(request):
	idTarifario		 = request.POST.get('idTarifario', "")
	idTarifaExtra    = request.POST.get('idTarifaExtra', "")
	nombre           = request.POST.get('extraName', "")

	if idTarifaExtra == "0":
		tarifaExtra = TarifaExtra()
		tarifaExtra.extra_descripcion = nombre
		tarifaExtra.tarifario = Tarifario.objects.get(id=idTarifario)
	else:
		tarifaExtra = TarifaExtra.objects.get(id=idTarifaExtra)

	for x in range(9):
		cat = x+1
		nameTramo='extra'+str(cat)
		valor = request.POST.get(nameTramo, "0")
		if valor == "":
			valor = 0
		if cat == 1:
			tarifaExtra.cat1 = float(valor)
		if cat == 2:
			tarifaExtra.cat2 = float(valor)
		if cat == 3:
			tarifaExtra.cat3 = float(valor)
		if cat == 4:
			tarifaExtra.cat4 = float(valor)
		if cat == 5:
			tarifaExtra.cat5 = float(valor)
		if cat == 6:
			tarifaExtra.cat6 = float(valor)
		if cat == 7:
			tarifaExtra.cat7 = float(valor)
		if cat == 8:
			tarifaExtra.cat8 = float(valor)
	tarifaExtra.save()

	tarifario = Tarifario.objects.get(id=idTarifario)
	context = {'tarifario': tarifario}
	return render(request, 'sistema/grillaTarifaExtra.html', context)

@login_required
def guardarMasivo(request):
	idTarifario = request.POST.get('idTarifario', "")
	tipoMasivo  = request.POST.get('tipoMasivo', "")
	porcentaje  = request.POST.get('porcentaje', "")

	tarifario = Tarifario.objects.get(id=idTarifario)

	if tipoMasivo == "1":
		for tv in tarifario.getTarifaViaje():
			for x in range(9):
				cat = x+1
				if cat == 1:
					tv.cat1 = int(tv.cat1) * int(float(porcentaje)) / 100 + int(tv.cat1)
				if cat == 2:
					tv.cat2 = int(tv.cat2) * int(float(porcentaje)) / 100 + int(tv.cat2)
				if cat == 3:
					tv.cat3 = int(tv.cat3) * int(float(porcentaje)) / 100 + int(tv.cat3)
				if cat == 4:
					tv.cat4 = int(tv.cat4) * int(float(porcentaje)) / 100 + int(tv.cat4)
				if cat == 5:
					tv.cat5 = int(tv.cat5) * int(float(porcentaje)) / 100 + int(tv.cat5)
				if cat == 6:
					tv.cat6 = int(tv.cat6) * int(float(porcentaje)) / 100 + int(tv.cat6)
				if cat == 7:
					tv.cat7 = int(tv.cat7) * int(float(porcentaje)) / 100 + int(tv.cat7)
				if cat == 8:
					tv.cat8 = int(tv.cat8) * int(float(porcentaje)) / 100 + int(tv.cat8)
			tv.save()
		nombre_html = "sistema/grillaTrayectoTarifa.html"


	if tipoMasivo == "2":
		for te in tarifario.getTarifaExtra():
			for x in range(9):
				cat = x+1
				if cat == 1:
					te.cat1 = int(te.cat1) * int(float(porcentaje)) / 100 + int(te.cat1)
				if cat == 2:
					te.cat2 = int(te.cat2) * int(float(porcentaje)) / 100 + int(te.cat2)
				if cat == 3:
					te.cat3 = int(te.cat3) * int(float(porcentaje)) / 100 + int(te.cat3)
				if cat == 4:
					te.cat4 = int(te.cat4) * int(float(porcentaje)) / 100 + int(te.cat4)
				if cat == 5:
					te.cat5 = int(te.cat5) * int(float(porcentaje)) / 100 + int(te.cat5)
				if cat == 6:
					te.cat6 = int(te.cat6) * int(float(porcentaje)) / 100 + int(te.cat6)
				if cat == 7:
					te.cat7 = int(te.cat7) * int(float(porcentaje)) / 100 + int(te.cat7)
				if cat == 8:
					te.cat8 = int(te.cat8) * int(float(porcentaje)) / 100 + int(te.cat8)
			te.save()

		nombre_html = "sistema/grillaTarifaExtra.html"

	context = {'tarifario': tarifario}
	return render(request, nombre_html, context)


@login_required
def listadoLicencia(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	licencias = Licencia.objects.all()

	context = {'licencias': licencias}
	return render(request, 'sistema/listadoLicencia.html', context)

@login_required
def altaLicencia(request):
	unidades = Unidad.objects.filter(baja=False)
	tipo_licencias = TipoLicencia.objects.all()
	licencia = Licencia()
	licencia.id = 0
	context = {'licencia': licencia, 'unidades': unidades, 'tipo_licencias':tipo_licencias}
	return render(request, 'sistema/licencia.html', context)

@login_required
def licencia(request):
	estado = request.session.get('estadoLicencia', '')
	idLicencia = request.GET.get('idLicencia', "")
	licencia = Licencia.objects.get(id=idLicencia)
	tipo_licencias = TipoLicencia.objects.all()
	unidades = Unidad.objects.filter(baja=False)
	request.session['estadoLicencia'] = ''
	context = {'tipo_licencias':tipo_licencias, 'estado':estado, 'licencia':licencia,'unidades':unidades}
	return render(request, 'sistema/licencia.html', context)

@login_required
def guardarLicencia(request):
	idLicencia = request.POST.get('idLicencia', False)
	tipoLicencia = request.POST.get('tipoLicencia', False)
	vencimientoLicencia = request.POST.get('vencimientoLicencia', False)
	descripcionLicencia = request.POST.get('descripcionLicencia', False)
	unidades = request.POST.get('unidades', False)

	if idLicencia == "0":
		licencia = Licencia()
		request.session['estadoLicencia'] = 'nuevo'
	else:
		licencia = Licencia.objects.get(id=idLicencia)
		request.session['estadoLicencia'] = 'editado'

	licencia.unidad_id = unidades
	licencia.comentario = descripcionLicencia
	licencia.tipo_licencia = TipoLicencia.objects.get(id=tipoLicencia)
	licencia.fecha_vencimiento = getAAAAMMDD(vencimientoLicencia)
	licencia.save()

	url = '/sistema/licencia/?idLicencia='+str(licencia.id)
	return redirect(url)

@login_required
def getSelectAsignoLicencia(request):
	listaAsignoLicencia = []
	tpl = request.POST.get('tipoPersonaLicencia', False)
	titulo = ""
	if tpl == "0":
		listaAsignoLicencia = Persona.objects.filter(tipo_persona=3,baja=False)
		titulo = "Chofer"
	elif tpl == "1":
		listaAsignoLicencia = Persona.objects.filter(tipo_persona=4,baja=False)
		titulo = "Dueño"
	elif tpl == "2":
		listaAsignoLicencia = Vehiculo.objects.all()
		titulo = "Vehículo"
	context = {'listaAsignoLicencia': listaAsignoLicencia, 'titulo':titulo}
	return render(request, 'sistema/selectAsignoLicencia.html', context)

@login_required
def cargarLocalidad(request):
	cp = request.POST.get('codigoPostal', False)
	localidades = Localidad.objects.filter(codigo_postal=cp)
	context = {'localidades': localidades}
	return render(request, 'sistema/selectLocalidad.html', context)

@login_required
def cargarLocalidadByDestino(request):
	destino_id = request.POST.get('destino_id', False)
	localidad_select_id = request.POST.get('localidad_select_id', '')
	localidades = Localidad.objects.filter(provincia_id=destino_id)
	context = {'localidades': localidades, 'localidad_select_id':localidad_select_id}
	return render(request, 'sistema/selectLocalidadViaje.html', context)

@login_required
def cargarProvincia(request):
	idLocalidad = request.POST.get('idLocalidad', False)
	localidad = Localidad.objects.get(id=idLocalidad)
	provincias = Provincia.objects.filter(id=localidad.provincia.id)
	context = {'provincias':provincias}
	return render(request, 'sistema/selectProvincia.html', context)

@login_required
def exportar(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	mensaje = ""

	context = {'mensaje': mensaje, 'clientes': Cliente.objects.filter(baja=False).values_list('id', 'razon_social'), 'unidades':Unidad.objects.filter(baja=False),'estados':Estado.objects.all()}
	return render(request, 'sistema/exportar.html', context)

@login_required
def cambiarEstadoViajes(request):
	idViajes = request.POST.get('idViajes', False)
	id_estado = request.POST.get('estadoModal', False)

	idsList = []
	for ids in idViajes.split("-"):
		if ids:
			idsList.append(int(ids))

	viajes = Viaje.objects.filter(id__in=idsList)
	for v in viajes:
		v.estado_id = id_estado
		v.save()

	retorno = 'Se actualizo el estado de los viajes seleccionados.'
	data = {'return': retorno}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

@login_required
def exportarDatosPorCliente(request):
	idCliente = request.GET.get('idCliente', False)

	cliente = Cliente.objects.get(id=idCliente)

	listadoPasajeros = []
	listadoSolicitantes = []
	for personacliente in cliente.personacliente_set.all():
		nombre_apellido = personacliente.persona.nombre + " " + personacliente.persona.apellido
		dict = {'id': personacliente.persona.id, 'valor': nombre_apellido }
		if personacliente.persona.tipo_persona.id == 1:
			listadoSolicitantes.append(dict)
		if personacliente.persona.tipo_persona.id == 2:
			listadoPasajeros.append(dict)

	listadoPasajeros = sorted(listadoPasajeros, key = lambda i: i['valor'])
	listadoSolicitantes = sorted(listadoSolicitantes, key = lambda i: i['valor'])

	listadoCC = []
	for cc in cliente.centrocosto_set.all():
		dict = {'id': cc.id, 'valor': cc.nombre }
		listadoCC.append(dict)

	listadoCC = sorted(listadoCC, key = lambda i: i['valor'])

	data = {'listadoPasajeros': listadoPasajeros, 'listadoSolicitantes': listadoSolicitantes, 'listadoCC': listadoCC}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

@login_required
def usuario(request):
	mensaje = ""

	context = {'mensaje': mensaje}
	return render(request, 'sistema/usuario.html', context)

@login_required
def listadoAdelanto(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	unidades = []
	permisos = obtenerPermiso(request)
	if 'unidades' in permisos:
		usrunidad = request.user.usrunidad_set.all()
		for u in usrunidad:
			unidades.append(u.unidad)
	else:
		unidades = Unidad.objects.filter(baja=False)

	mensaje = ""
	context = {'mensaje': mensaje, 'unidades':unidades}
	return render(request, 'sistema/listadoAdelanto.html', context)

@login_required
def altaAdelanto(request):
	mensaje = ""

	unidades = []
	permisos = obtenerPermiso(request)
	if 'unidades' in permisos:
		usrunidad = request.user.usrunidad_set.all()
		for u in usrunidad:
			unidades.append(u.unidad)
	else:
		unidades = Unidad.objects.filter(baja=False)

	tipos_adelanto = TipoAdelanto.objects.all()
	adelanto = Adelanto()
	adelanto.id = 0
	context = {'mensaje': mensaje, 'unidades':unidades,'tipos_adelanto':tipos_adelanto,'adelanto':adelanto}
	return render(request, 'sistema/adelanto.html', context)

@login_required
def adelanto(request):
	mensaje = ""

	estado = request.session.get('estadoAdelanto', '')
	idAdelanto = request.GET.get('idAdelanto', "")
	adelanto = Adelanto.objects.get(id=idAdelanto)
	tipos_adelanto = TipoAdelanto.objects.all()

	unidades = []
	permisos = obtenerPermiso(request)
	if 'unidades' in permisos:
		usrunidad = request.user.usrunidad_set.all()
		for u in usrunidad:
			unidades.append(u.unidad)
	else:
		unidades = Unidad.objects.filter(baja=False)

	tipoAdelantoId = adelanto.tipo_adelanto.id
	unidadId = adelanto.unidad.id
	request.session['estadoAdelanto'] = ''

	context = {'mensaje': mensaje,
				'unidades':unidades,
				'tipos_adelanto':tipos_adelanto,
				'estado':estado,
				'idAdelanto':idAdelanto,
				'tipoAdelantoId':tipoAdelantoId,
				'adelanto':adelanto,
				'unidadId':unidadId
			}
	return render(request, 'sistema/adelanto.html', context)

@login_required
def guardarAdelanto(request):
	mensaje = ""
	idAdelanto = request.POST.get('idAdelanto', False)
	unidad = request.POST.get('unidad', False)
	monto = request.POST.get('monto', False)
	tipoAdelanto = request.POST.get('tipoAdelanto', False)
	descripcion = request.POST.get('descripcion', False)
	fecha = request.POST.get('fecha', False)
	factura = request.POST.get('factura', False)

	if idAdelanto == "0":
		adelanto = Adelanto()
		request.session['estadoAdelanto'] = 'nuevo'
	else:
		adelanto = Adelanto.objects.get(id=idAdelanto)
		request.session['estadoAdelanto'] = 'editado'

	adelanto.unidad = Unidad.objects.get(id=unidad)
	adelanto.monto = monto
	adelanto.tipo_adelanto = TipoAdelanto.objects.get(id=tipoAdelanto)
	adelanto.descripcion = descripcion
	adelanto.fecha = getAAAAMMDD(fecha)
	adelanto.factura = factura
	adelanto.save()

	url = '/sistema/adelanto/?idAdelanto='+str(adelanto.id)
	return redirect(url)

@login_required
def buscarAdelantos(request):

	fechaDesde = request.POST.get('desde', False)
	fechaHasta = request.POST.get('hasta', False)
	unidad = request.POST.get('unidad', False)

	fechaDesde =  getAAAAMMDD(fechaDesde)
	fechaHasta =  getAAAAMMDD(fechaHasta)

	adelantos = []
	if unidad:
		adelantos = Adelanto.objects.filter(unidad_id=unidad,fecha__gte=fechaDesde, fecha__lte=fechaHasta)
	else:
		adelantos = Adelanto.objects.filter(fecha__gte=fechaDesde, fecha__lte=fechaHasta)

	context = {'adelantos': adelantos}
	return render(request, 'sistema/grillaAdelantos.html', context)

@login_required
def eliminarAdelanto(request):
	idAdelanto = request.GET.get('idAdelanto', False)
	Adelanto.objects.get(id=idAdelanto).delete()
	data = {'return': 'success'}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')


@login_required
def facturarAdelantos(request):
	idAdelantos = request.POST.get('idAdelantos', False)
	numeroFactura = request.POST.get('numeroFactura', False)
	idsList = []
	for ids in idAdelantos.split("-"):
		if ids:
			idsList.append(int(ids))

	adelantos = Adelanto.objects.filter(id__in=idsList)
	for a in adelantos:
		a.factura = numeroFactura
		a.save()

	data = {'return': 'success'}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

@login_required
def listadoFactClientes(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	clientes = Cliente.objects.filter(baja=False)
	categorias = CategoriaViaje.objects.all()
	condicionesPago = CondicionPago.objects.all()
	estados = Estado.objects.all()
	cc = CentroCosto.objects.filter(baja=False)
	context = {'clientes': clientes, 'categorias':categorias, 'condicionesPago':condicionesPago, 'estados':estados, 'centrosCosto': cc}
	return render(request, 'sistema/listadoFactClientes.html', context)

@login_required
def cargarCentrosDeCosto(request):
	idCliente = request.GET.get('idCliente', False)
	centrosDeCosto = []
	if idCliente:
		centrosDeCosto = CentroCosto.objects.filter(cliente_id=idCliente)

	context = {'centrosDeCosto': centrosDeCosto}
	return render(request, 'sistema/selectCentroCostos.html', context)

@login_required
def cargarSolicitantes(request):
	idCliente = request.GET.get('idCliente', False)
	solicitantes = []
	if idCliente:
		solicitantes = Persona.objects.filter(personacliente__cliente_id=idCliente, tipo_persona_id=1)

	context = {'solicitantes': solicitantes}
	return render(request, 'sistema/selectSolicitantes.html', context)

@login_required
def cargarFactura(request):
	idCliente = request.GET.get('idCliente', False)
	facturas = []
	if idCliente:
		viajes = FacturaViaje.objects.filter(viaje__cliente_id=idCliente).order_by('fact_cliente')
		for v in viajes:
			if v.fact_cliente not in facturas:
				if v.fact_cliente:
					facturas.append(v.fact_cliente)

	context = {'facturas': facturas}
	return render(request, 'sistema/selectFacturas.html', context)

@login_required
def cargarFacturaUnidad(request):
	idUnidad = request.GET.get('idUnidad', False)
	uniList = []
	if idUnidad != "null":
		for c in idUnidad.split(","):
			uniList.append(int(c))

	facturas = []
	if uniList:
		viajes = FacturaViaje.objects.filter(viaje__unidad_id__in=uniList).order_by('fact_proveedor')
		for v in viajes:
			if v.fact_proveedor not in facturas:
				if v.fact_proveedor:
					facturas.append(v.fact_proveedor)

	context = {'facturas': facturas}
	return render(request, 'sistema/selectFacturas.html', context)

@login_required
def cargarProforma(request):
	idCliente = request.GET.get('idCliente', False)
	proformas = []
	if idCliente:
		viajes = FacturaViaje.objects.filter(viaje__cliente_id=idCliente).order_by('prof_cliente')
		for v in viajes:
			if v.prof_cliente in proformas:
				pass
			else:
				if v.prof_cliente:
					proformas.append(v.prof_cliente)

	context = {'proformas': proformas}
	return render(request, 'sistema/selectProformas.html', context)

@login_required
def buscarFacturacionCliente(request):
	idCliente 		= request.POST.get('cliente', False)
	categorias 		= request.POST.get('categorias', False)
	estados 		= request.POST.get('estados', False)
	centroDeCosto 	= request.POST.get('centroDeCosto', False)
	condEspecial 	= request.POST.get('condEspecial', False)
	solicitantes 	= request.POST.get('solicitantes', False)
	facturas 		= request.POST.get('facturas', False)
	proformas 		= request.POST.get('proformas', False)
	desde 			= request.POST.get('desde', False)
	hasta 			= request.POST.get('hasta', False)

	fechaDesde =  getAAAAMMDD(desde)
	fechaHasta =  getAAAAMMDD(hasta)

	listaviajes = []

	catList = []
	if categorias != "null":
		for c in categorias.split(","):
			catList.append(int(c))

	estList = []
	if estados != "null":
		for c in estados.split(","):
			estList.append(int(c))

	ccList = []
	if centroDeCosto != "null":
		for c in centroDeCosto.split(","):
			ccList.append(int(c))

	facList = []
	sinFactura = False
	if facturas != "null":
		for c in facturas.split(","):
			if c == "0":
				sinFactura = True
			facList.append(c)

	proList = []
	sinProforma = False
	if proformas != "null":
		for c in proformas.split(","):
			if c == "0":
				sinProforma = True
			proList.append(c)

	solList = []
	if solicitantes != "null":
		for c in solicitantes.split(","):
			solList.append(int(c))

	if idCliente:
		viajes = Viaje.objects.filter(cliente_id=idCliente,fecha__gte=fechaDesde, fecha__lte=fechaHasta)
	else:
		viajes = Viaje.objects.filter(fecha__gte=fechaDesde, fecha__lte=fechaHasta)

	if catList:
		viajes = viajes.filter(categoria_viaje_id__in=catList)

	if ccList:
		viajes = viajes.filter(centro_costo_id__in=ccList)

	if solList:
		viajes = viajes.filter(solicitante_id__in=solList)

	if condEspecial == "1":
		q_ids = [o.id for o in viajes if o.getMontoEstacionCliente() == 0]
		viajes = viajes.filter(id__in=q_ids)
	elif condEspecial == "2":
		q_ids = [o.id for o in viajes if o.getMontoEstacionCliente() > 0]
		viajes = viajes.filter(id__in=q_ids)
	elif condEspecial == "3":
		viajes = viajes.filter(categoria_viaje_id__in=[1,2,3,4,17,18,19])
	elif condEspecial == "4":
		viajes = viajes.filter(categoria_viaje_id__in=[5,6,7,8,9])
	elif condEspecial == "5":
		viajes = viajes.filter(estado_id__in=[1,2,3,4,5,6,9,10])

	if estList:
		viajes = viajes.filter(estado_id__in=estList)

	if sinProforma or sinFactura:
		if sinProforma:
			q_ids = [o.id for o in viajes if o.getProforma() == ""]
			viajes = viajes.filter(id__in=q_ids)
		if sinFactura:
			q_ids = [o.id for o in viajes if o.getFacturaCliente() == ""]
			viajes = viajes.filter(id__in=q_ids)
	else:
		if facList:
			q_ids = [o.id for o in viajes if o.getFacturaCliente() in facList]
			viajes = viajes.filter(id__in=q_ids)
		if proList:
			q_ids = [o.id for o in viajes if o.getProforma() in proList]
			viajes = viajes.filter(id__in=q_ids)

	context = {'viajes': viajes}
	return render(request, 'sistema/grillaFacturacionCliente.html', context)

@login_required
def buscarFacturacionProveedor(request):
	idUnidad 		= request.POST.get('unidad', False)
	condEspecial 	= request.POST.get('condEspecial', False)
	facturas 		= request.POST.get('facturas', False)
	proveedor 		= request.POST.get('proveedor', False)
	estados 		= request.POST.get('estados', False)
	desde 			= request.POST.get('desde', False)
	hasta 			= request.POST.get('hasta', False)

	fechaDesde =  getAAAAMMDD(desde)
	fechaHasta =  getAAAAMMDD(hasta)

	listaviajes = []

	facList = []
	sinFactura = False
	if facturas != "null":
		for c in facturas.split(","):
			if c == "0":
				sinFactura = True
			facList.append(c)

	estList = []
	if estados != "null":
		for c in estados.split(","):
			estList.append(int(c))

	uniList = []
	if idUnidad != "null":
		for c in idUnidad.split(","):
			uniList.append(int(c))

	if sinFactura:
		if uniList:
			viajes = Viaje.objects.filter(unidad_id__in=uniList,fecha__gte=fechaDesde, fecha__lte=fechaHasta)
		else:
			viajes = Viaje.objects.filter(fecha__gte=fechaDesde, fecha__lte=fechaHasta)
		q_ids = [o.id for o in viajes if o.getFacturaProveedor()==""]
		viajes = viajes.filter(id__in=q_ids)

	else:
		viajes = Viaje.objects.filter(unidad_id__in=uniList,fecha__gte=fechaDesde, fecha__lte=fechaHasta)
		if facList:
			q_ids = [o.id for o in viajes if o.getFacturaProveedor() in facList]
			viajes = viajes.filter(id__in=q_ids)

	if estList:
		viajes = viajes.filter(estado_id__in=estList)

	context = {'viajes': viajes}
	return render(request, 'sistema/grillaFacturacionProveedor.html', context)

@login_required
def facturarClientes(request):
	idViajes 			= request.POST.get('idViajes', False)
	numeroFactrura		= request.POST.get('numeroFactura', False)
	primernumeroFactura = request.POST.get('primernumeroFactura', False)
	idsList = []
	for ids in idViajes.split("-"):
		if ids:
			idsList.append(int(ids))

	viajes = Viaje.objects.filter(id__in=idsList)
	for v in viajes:
		if v.facturaviaje_set.all():
			fv = v.facturaviaje_set.all()[0]
		else:
			fv = FacturaViaje()
			fv.viaje = v

		fv.fact_cliente = primernumeroFactura + "-" + numeroFactrura
		fv.save()

	data = {'return': 'success'}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

@login_required
def facturarProveedores(request):
	idViajes 			= request.POST.get('idViajes', False)
	numeroFactrura		= request.POST.get('numeroFactura', False)
	primernumeroFactura = request.POST.get('primernumeroFactura', False)
	idsList = []
	for ids in idViajes.split("-"):
		if ids:
			idsList.append(int(ids))

	viajes = Viaje.objects.filter(id__in=idsList)
	for v in viajes:
		if v.facturaviaje_set.all():
			fv = v.facturaviaje_set.all()[0]
		else:
			fv = FacturaViaje()
			fv.viaje = v

		fv.fact_proveedor = primernumeroFactura + "-" + numeroFactrura
		fv.save()

	data = {'return': 'success'}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

@login_required
def proformarClientes(request):
	idViajes = request.POST.get('idViajes', False)

	idsList = []
	for ids in idViajes.split("-"):
		if ids:
			idsList.append(int(ids))

	viajes = Viaje.objects.filter(id__in=idsList)

	estadosCorrectos = True
	for v in viajes:
		if v.estado.id != 7:
			estadosCorrectos = False

	if not estadosCorrectos:
		retorno = 'Alguno de los viajes seleccionados no esta cerrado.'
		data = {'return': retorno}
		dump = json.dumps(data)
		return HttpResponse(dump, content_type='application/json')

	facturasViaje = FacturaViaje.objects.all().order_by('-prof_cliente')

	fvs = FacturaViaje.objects.exclude(prof_cliente__isnull=True).exclude(prof_cliente__exact='').values_list('prof_cliente', flat=True)
	fvs = list(fvs)
	numbers = [ int(x) for x in fvs ]
	print max(numbers)

	if fvs:
		numeroProforma = max(numbers) + 1
	else:
		numeroProforma = 1

	proformados = ''
	for v in viajes:
		if v.facturaviaje_set.all():
			fv = v.facturaviaje_set.all()[0]
		else:
			fv = FacturaViaje()
			fv.viaje = v

		fv.prof_cliente = numeroProforma
		fv.save()
		proformados += str(v.id) + ','

	retorno = 'Viajes con numero de proforma ' + str(numeroProforma) + ': ' + proformados
	data = {'return': retorno}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

@login_required
def listadoFactProvedores(request):
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	unidades = Unidad.objects.all()
	estados = Estado.objects.all()
	context = {'unidades': unidades,'estados':estados}
	return render(request, 'sistema/listadoFactProvedores.html', context)

@login_required
def cargarFacturaProveedor(request):
	idUnidad = request.GET.get('idUnidad', False)
	facturas = []
	if idUnidad:
		viajes = FacturaViaje.objects.filter(viaje__unidad_id=idUnidad).order_by('fact_proveedor')
		for v in viajes:
			if v.fact_proveedor not in facturas:
				if v.fact_proveedor:
					facturas.append(v.fact_proveedor)

	context = {'facturas': facturas}
	return render(request, 'sistema/selectFacturas.html', context)

@login_required
def borrarSolicitanteCliente(request):
	idPersona = request.POST.get('idPersona', False)
	idCliente = request.POST.get('idCliente', False)

	persona = Persona.objects.get(id=idPersona)
	persona.baja = True
	persona.save()

	cliente = Cliente.objects.get(id=idCliente)
	for c in cliente.personacliente_set.all():
		if c.persona.tipo_persona.id == 1 and c.persona.id == persona.id:
			c.delete()

	context = {'cliente':cliente}
	return render(request, 'sistema/grillaSolicitantes.html', context)

@login_required
def borrarPasajeroCliente(request):
	idPersona = request.POST.get('idPersona', False)
	idCliente = request.POST.get('idCliente', False)

	persona = Persona.objects.get(id=idPersona)
	persona.baja = True
	persona.save()

	cliente = Cliente.objects.get(id=idCliente)
	for c in cliente.personacliente_set.all():
		if c.persona.tipo_persona.id == 2 and c.persona.id == persona.id:
			c.delete()

	cliente_id = cliente.id
	pasajeros = cliente.getPasajeros()

	context = {'cliente_id':cliente_id, 'pasajeros':pasajeros}
	return render(request, 'sistema/grillaPasajeros.html', context)

@login_required
def borrarCCPropect(request):
	idCC = request.POST.get('idCC', False)
	idCliente = request.POST.get('idCliente', False)

	cc = CentroCosto.objects.get(id=idCC)
	cc.baja = True
	cc.save()

	cliente = Cliente.objects.get(id=idCliente)

	context = {'cliente':cliente }
	return render(request, 'sistema/grillaCentroCostos.html', context)

@login_required
def exportarExcelFactCliente(request):
	ids = request.POST.get('ids', False)

	from openpyxl import Workbook
	wb = Workbook()

	# grab the active worksheet
	ws = wb.active

	# Data can be assigned directly to cells
	ws['A1'] = 42

	# Rows can also be appended
	ws.append([1, 2, 3])

	# Python types will automatically be converted
	ws['A2'] = datetime.datetime.now()

	# Save the file
	wb.save('/static/excel_fact_clie/excel_fact_cliente.xls')

	file_name = '/static/excel_fact_clie/excel_fact_cliente.xls'
	data = {'file_name': file_name}
	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')

@login_required
def exportarPdfFactCliente(request):
	cliente 	= request.GET['cliente']
	desde   	= request.GET['desde']
	hasta   	= request.GET['hasta']
	idsViaje	= request.GET['ids']
	centroCosto	= request.GET['centrosCosto']

	retornoCC = ""
	ccList = []
	if centroCosto != "null":
		for c in centroCosto.split(","):
			ccList.append(int(c))
	if ccList:
		cc = CentroCosto.objects.filter(id__in=ccList)
		for c in cc:
			aux = str(c.id) +"-"+ c.nombre
			retornoCC = retornoCC + aux + " / "


	cliente = Cliente.objects.get(id=cliente)
	idsList = []
	for ids in idsViaje.split("-"):
		if ids:
			idsList.append(int(ids))

	hsdispo = 0
	dispo = 0
	subtotal = 0
	peft = 0
	tiempo = 0
	mtiempo = 0
	bilingue = 0
	monto = 0
	peaje = 0
	estacion = 0
	otros = 0
	total = 0
	iva   = 0
	final = 0
	viajes = Viaje.objects.filter(id__in=idsList).order_by('fecha', 'hora')
	for v in viajes:
		total = total + v.getTotalCliente()
		iva   = iva + v.getIvaCliente()
		final = final + v.getFinalCliente()
		hsdispo = hsdispo + v.getHsDispoCliente()
		dispo = dispo + v.getMontoDispoCliente()
		subtotal = subtotal + v.getSubtotalCliente()
		peft = peft + v.getMontoPeftCliente()
		mtiempo = mtiempo + v.getMontoTiempoEsperaCliente()
		tiempo = tiempo + v.getCantidadTiempoEsperaCliente()
		bilingue = bilingue + v.getMontoBilingueCliente()
		monto = monto + v.getMontoMontoCliente()
		peaje = peaje + v.getMontoPeajesCliente()
		estacion = estacion + v.getMontoEstacionCliente()
		otros = otros + v.getMontoOtrosCliente()

	context = {'cliente': cliente, 'desde':desde, 'hasta': hasta, 'viajes':viajes, 'total':total, 'iva': iva, 'final': final, 'subtotal':subtotal,'peft':peft,'tiempo':tiempo,'mtiempo':mtiempo, 'bilingue':bilingue,'monto':monto,'peaje':peaje,'estacion':estacion,'otros':otros, 'centroCosto':retornoCC, 'hsdispo':hsdispo, 'dispo': dispo}
	return render(request, 'sistema/pdfFactCliente.html', context)

@login_required
def exportarPdfFactProv(request):
	idunidad 	= request.GET['unidad']
	desde   	= request.GET['desde']
	hasta   	= request.GET['hasta']
	idsViaje	= request.GET['ids']

	unidades = ""
	for ids in idunidad.split(","):
		if ids:
			prov = Unidad.objects.get(id=ids)
			unidades = unidades + prov.id_fake + " - " + prov.identificacion + " // "



	idsList = []
	for ids in idsViaje.split("-"):
		if ids:
			idsList.append(int(ids))

	subtotal = 0
	hsdispo = 0
	dispo = 0
	cobrado = 0
	tiempo = 0
	mtiempo = 0
	bilingue = 0
	maletas = 0
	peajes = 0
	estacion = 0
	otros = 0
	total = 0
	iva = 0
	final = 0
	pagar = 0

	viajes = Viaje.objects.filter(id__in=idsList).order_by('fecha', 'hora')
	for v in viajes:
		subtotal = subtotal + v.getSubtotalProveedor()
		hsdispo = hsdispo + v.getHsDispoProveedor()
		dispo = dispo + v.getMontoDispoProveedor()
		cobrado = cobrado + v.getCobradoProveedor()
		tiempo = tiempo + v.getCantidadTiempoEsperaProveedor()
		mtiempo = mtiempo + v.getMontoTiempoEsperaProveedor()
		bilingue = bilingue + v.getMontoBilingueProveedor()
		maletas = maletas + v.getMontoMaletasProveedor()
		peajes = peajes + v.getMontoPeajesProveedor()
		estacion = estacion + v.getMontoEstacionProveedor()
		otros = otros + v.getMontoOtrosProveedor()
		total = total + v.getTotalProveedor()
		iva = iva + v.getIvaProveedor()
		final = final + v.getFinalProveedor()
		pagar = pagar + v.getPagarProveedor()

	context = {'unidades': unidades, 'desde':desde, 'hasta': hasta, 'viajes':viajes, 'subtotal':subtotal,'hsdispo':hsdispo,'dispo':dispo,'cobrado':cobrado,'tiempo':tiempo,'mtiempo':mtiempo,'bilingue':bilingue,'maletas':maletas,'peajes':peajes,'estacion':estacion,'otros':otros,'total':total,'iva':iva,'final':final,'pagar':pagar}
	return render(request, 'sistema/pdfFactProvedor.html', context)

@login_required
def exportarPdfViaje(request):
	mensaje = ""
	idViaje = request.GET['idViaje']
	viaje = Viaje.objects.get(id=idViaje)
	context = {'viaje' : viaje}
	return render(request, 'sistema/pdfViaje.html', context)

@login_required
def mailtoViaje(request):
	idViaje = request.POST.get('idViaje', False)
	viaje = Viaje.objects.get(id=idViaje)

	body = '''Estimados,%0ALes informamos el estado del servicio del pasajero de referencia:%0DServicio:	{} %0D%0A
	Fecha: {} %0D%0A
	Hora: {}hs %0D%0A
	Desde: {} %0D%0A
	Hasta: {} %0D%0A
	C. Costos: {} %0D%0A
	Nota: %0D%0A %0D%0A %0D%0A
	Cordialmente, %0D%0A
	Reservas | LOGOSTRASLADOS | T +5411 5031-3800 | WhatsApp +54911 3193-1428 | M: reservas@logostraslados.com.ar | W: www.logostraslados.com.ar
	'''.format(viaje.id,viaje.getFecha(), viaje.getHora(), viaje.getTrayectoPrincipal().desdeConcat() if viaje.getTrayectoPrincipal() else '', viaje.getTrayectoPrincipal().hastaConcat() if viaje.getTrayectoPrincipal() else '', viaje.centro_costo.nombre)

	data = {
		'mailto': '%s' %(viaje.solicitante.getMail()),
		'subject': 'Informe | %s | Pax: %s %s' %(viaje.cliente.razon_social, viaje.pasajero.apellido, viaje.pasajero.nombre),
		'mailtocco': 'informes@logostraslados.com.ar',
		'body': body
	}

	dump = json.dumps(data)
	return HttpResponse(dump, content_type='application/json')


@login_required
def unidadViaje(request):
   
	if not validarUrlPorRol(request):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	id_viaje = request.GET.get('idViaje', '')
	viaje = Viaje.objects.get(id=id_viaje)

	if not validarViajeUsuarioUnidad(request, viaje):
		mensaje = ""
		context = { 'mensaje':mensaje }
		return render(request, 'sistema/urlBloqueada.html', context)

	context = {
		'id': id_viaje,
        'viaje': viaje,
		'fecha' : viaje.getFecha(),
		'hora' : viaje.hora,
		#'estimados' : viaje.hora_estimada,
		'estado' : viaje.estado.estado,
		'categoria_viaje' : viaje.categoria_viaje.categoria,
		'comentario' : viaje.getObservacioneChofer(),
		'pasajero' : viaje.pasajero.nombreCompleto(),
		'cantidad' : viaje.nropasajeros,
		'cliente' : viaje.cliente.razon_social,
		'espera' : str(timedelta(minutes=int(viaje.espera)))[:-3] + 'hs' if viaje.espera else '',
		'hs_dispo' : viaje.dispo,
		'bilingue' : viaje.bilingue,
		'maletas' : viaje.maletas,
		'peajes' : viaje.peajes,
		'estacionamiento' : viaje.parking,
		'otros' : viaje.otro,
		'desde' : viaje.getTrayectoPrincipal().desdeConcat(),
		'hasta' : viaje.getTrayectoPrincipal().hastaConcat(),
		'trayectos' : Trayecto.objects.filter(viaje_id=id_viaje),
		'adjuntos' : viaje.getViajeAdjuntos()
	}

	return render(request, 'sistema/unidadViaje.html', context)



@login_required
def unidadDashboard(request):
	mensaje = ""
	context = {'mensaje': mensaje}
	return render(request, 'sistema/unidadDashboard.html', context)



@login_required
def refreshUnidadDashboard(request):
	desde = request.POST.get('desde', False)
	hasta = request.POST.get('hasta', False)

	fechaDesde =  getAAAAMMDD(desde)
	fechaHasta =  getAAAAMMDD(hasta)
	uniList    = []

	current_user = request.user

	uni_usu = current_user.usrunidad_set.all()
	unidades = []
	for uu in uni_usu:
		if uu.unidad:
			unidades.append(uu.unidad.id)

	if unidades:
		viajes = Viaje.objects.filter(unidad_id__in=unidades,fecha__gte=fechaDesde, fecha__lte=fechaHasta)
	else:
		viajes = []

	recaudado   = 0
	promedio_dia = 0
	promedio_viaje = 0
	adelantos_valor = 0
	cant_viajes = 0
	dias_trab   = []

	adelantos = Adelanto.objects.filter(unidad_id__in=unidades,fecha__gte=fechaDesde, fecha__lte=fechaHasta)

	for a in adelantos:
		adelantos_valor += a.monto

	for v in viajes:
		if v.fecha not in dias_trab:
			dias_trab.append(v.fecha)
		cant_viajes += 1 
		recaudado += v.getPagarProveedor()

	try:
		promedio_dia = round(recaudado/len(dias_trab))
		promedio_viaje = round(recaudado/cant_viajes)
	except Exception as exception:
		promedio_dia = 0
		promedio_viaje = 0

	context = {
		'recaudado': recaudado,
		'promedio_dia' : promedio_dia,
		'promedio_viaje' : promedio_viaje,
		'adelantos': adelantos_valor,
		'cantidad_viajes': cant_viajes,
		'dias_trabajados': len(dias_trab)
	}
	return render(request, 'sistema/unidadDashboardDetails.html', context)


@login_required
def cargarMenu(request):
	permisos = [x.name for x in Permission.objects.filter(user=request.user)]
	if 'unidades' in permisos:
		menu_file = "sistema/unidadesMenu.html"
	if 'operaciones' in permisos:
		menu_file = "sistema/operacionesMenu.html"
	if 'finanzas' in permisos:
		menu_file = "sistema/finanzasMenu.html"
	if 'superuser' in permisos:
		menu_file = "sistema/superUserMenu.html"

	context = {'facturas': ''}
	return render(request, menu_file, context)

@login_required
def urlBloqueada(request):
    mensaje = ""

    context = {'mensaje': mensaje}
    return render(request, 'sistema/urlBloqueada.html', context)

# devuelve AAAAMMDD
def fecha():
	import time
	return time.strftime("%Y%m%d%H%M")

def getAAAAMMDD(fecha):
	return fecha[6:10] + fecha[3:5] + fecha[0:2]

@register.filter
def get_tarifa_by_cat(tarifaTrayecto, idCat):
    return tarifaTrayecto.getTarifaByCategoria(idCat)

@register.filter
def get_tarifa_extra_by_cat(tarifaExtra, idCat):
    return tarifaExtra.getTarifaExtraByCategoria(idCat)

def obtenerPermiso(request):
	permisos = [x.name for x in Permission.objects.filter(user=request.user)]
	menu_file = ""
	if 'unidades' in permisos:
		menu_file = 'unidades'
	if 'operaciones' in permisos:
		menu_file = 'operaciones'
	if 'finanzas' in permisos:
		menu_file = 'finanzas'
	if 'superuser' in permisos:
		menu_file = 'superuser'

	return menu_file

def validarUrlPorRol(request):
	permisos = obtenerPermiso(request)
	if 'unidades' in permisos:
		urls = ['asignaciones','listadoAdelanto','listadoFactProvedores','password_change','unidadViaje']
	if 'operaciones' in permisos:
		urls = ['operaciones','editaViaje','altaViaje','exportar','listadoCliente','listadoCentroDeCosto','listadoTarifario','listadoContacto','listadoProvedor','listadoUnidad','listadoLicencia','password_change','editaViaje']
	if 'finanzas' in permisos:
		urls = ['operaciones','altaViaje','editaViaje','exportar','listadoCliente','listadoCentroDeCosto','listadoTarifario','listadoContacto','listadoProvedor','listadoUnidad','listadoLicencia','listadoAdelanto','listadoFactClientes','listadoFactProvedores','password_change','editaViaje']
	if 'superuser' in permisos:
		urls = ['operaciones','altaViaje','exportar','editaViaje','asignaciones','listadoCliente','listadoCentroDeCosto','listadoTarifario','listadoContacto','listadoProvedor','listadoUnidad','listadoLicencia','listadoAdelanto','listadoFactClientes','listadoFactProvedores','password_change','editaViaje','unidadViaje']

	for url in urls:
		if url in request.build_absolute_uri():
			return True
	return False

def validarViajeUsuarioUnidad(request, viaje):
	if validaViajeUnidad(request, viaje):
		return viaje
	else:
		return False

def validaViajeUnidad(request, viaje):
	unidades = getUnidadesByUser(request)
	permisos = obtenerPermiso(request)
	if 'unidades' in permisos:
		if viaje.unidad:
			if viaje.unidad.id in unidades:
				return viaje
			else:
				return False
	return viaje

def getUnidadesByUser(request):
	usrunidad = request.user.usrunidad_set.all()
	unidades = []

	for u in usrunidad:
		unidades.append(u.unidad.id)
	return unidades